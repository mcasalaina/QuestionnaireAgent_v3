"""Excel file loading and saving functionality."""

import openpyxl
from typing import List, Optional
from utils.data_types import WorkbookData, SheetData, CellState
from utils.exceptions import ExcelFormatError
import logging
import os

logger = logging.getLogger(__name__)


class ExcelLoader:
    """Loads and saves Excel workbooks for live processing visualization."""
    
    def __init__(self, column_identifier=None):
        """Initialize Excel loader.

        Args:
            column_identifier: Optional ColumnIdentifier instance for detecting columns.
                             If None, will use fallback (column A = Question, B = Response).
        """
        self.column_identifier = column_identifier

    def _run_async_in_thread(self, coro):
        """Run an async coroutine in a separate thread with its own event loop.

        This is necessary when calling async code from a context that already
        has a running event loop (like FastAPI).

        Args:
            coro: The coroutine to run

        Returns:
            The result of the coroutine
        """
        import asyncio
        import threading

        result = None
        error = None

        def run_in_thread():
            nonlocal result, error
            try:
                # Create a new event loop for this thread
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    result = loop.run_until_complete(coro)
                finally:
                    # Clean up pending tasks before closing the loop
                    pending = asyncio.all_tasks(loop)
                    for task in pending:
                        task.cancel()
                    # Give tasks a chance to complete cancellation
                    if pending:
                        loop.run_until_complete(asyncio.gather(*pending, return_exceptions=True))
                    loop.close()
            except Exception as e:
                error = e

        thread = threading.Thread(target=run_in_thread)
        thread.start()
        thread.join(timeout=60)  # Add timeout to prevent hanging

        if thread.is_alive():
            raise TimeoutError("Async operation timed out after 60 seconds")

        if error:
            raise error
        return result

    def _run_async_in_thread_v2(self, async_func):
        """Run an async function in a separate thread with its own event loop.

        This version takes an async function (not a coroutine), so the coroutine
        is created inside the thread's event loop, avoiding issues with event loop
        references.

        Args:
            async_func: An async function to call (not a coroutine)

        Returns:
            The result of the async function
        """
        import asyncio
        import threading

        result = None
        error = None

        def run_in_thread():
            nonlocal result, error
            try:
                # Create a new event loop for this thread
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    # Call the async function inside this thread's event loop
                    result = loop.run_until_complete(async_func())
                finally:
                    # Clean up pending tasks before closing the loop
                    pending = asyncio.all_tasks(loop)
                    for task in pending:
                        task.cancel()
                    # Give tasks a chance to complete cancellation
                    if pending:
                        loop.run_until_complete(asyncio.gather(*pending, return_exceptions=True))
                    loop.close()
            except Exception as e:
                error = e

        thread = threading.Thread(target=run_in_thread)
        thread.start()
        thread.join(timeout=60)  # Add timeout to prevent hanging

        if thread.is_alive():
            raise TimeoutError("Async operation timed out after 60 seconds")

        if error:
            raise error
        return result
    
    def load_workbook(self, file_path: str) -> WorkbookData:
        """Load Excel file into WorkbookData structure.
        
        Args:
            file_path: Absolute path to Excel file (.xlsx or .xls)
            
        Returns:
            WorkbookData with all visible sheets and questions
            
        Raises:
            FileNotFoundError: If file doesn't exist
            ExcelFormatError: If file format is invalid
            ExcelFormatError: If no visible sheets found
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        if not file_path.lower().endswith(('.xlsx', '.xls')):
            raise ExcelFormatError("File must be an Excel file (.xlsx or .xls)")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ExcelFormatError(f"Invalid Excel file: {e}")
        
        sheets = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Skip hidden sheets
            if ws.sheet_state != 'visible':
                logger.info(f"Skipping hidden sheet: {sheet_name}")
                continue

            # Find the actual header row (it may not be row 1)
            header_row_num = self._find_header_row(ws)
            if header_row_num is None:
                logger.warning(f"Sheet '{sheet_name}' has no recognizable header row, skipping")
                continue

            # Get headers from identified header row
            headers = []
            for cell in ws[header_row_num]:
                headers.append(cell.value if cell.value else '')

            if not headers:
                logger.warning(f"Sheet '{sheet_name}' has no headers, skipping")
                continue

            logger.info(f"Found headers in row {header_row_num} for sheet '{sheet_name}'")
            
            # Identify columns using column identifier or fallback to defaults
            column_mapping = {'question': 0, 'response': 1, 'documentation': None}
            if self.column_identifier:
                try:
                    # Column identification is now synchronous
                    column_mapping = self.column_identifier.identify_columns(headers)
                    logger.info(f"Identified columns for sheet '{sheet_name}': {column_mapping}")
                except Exception as e:
                    logger.warning(f"Failed to identify columns, using defaults (A=Question, B=Response): {e}")
            
            question_col = column_mapping.get('question')
            response_col = column_mapping.get('response')
            doc_col = column_mapping.get('documentation')
            
            if question_col is None:
                logger.warning(f"Sheet '{sheet_name}' has no question column identified, skipping")
                continue
            
            # Extract questions from identified question column (0-based to 1-based for openpyxl)
            # Collect all non-blank rows first, then use LLM to filter
            all_rows = []
            for row in ws.iter_rows(min_row=header_row_num+1, min_col=question_col+1, max_col=question_col+1):
                cell = row[0]
                cell_value = str(cell.value).strip() if cell.value else ''

                # Skip blank cells
                if not cell_value:
                    logger.debug(f"Skipping blank row {cell.row} in sheet '{sheet_name}'")
                    continue

                all_rows.append({
                    'row_num': cell.row,
                    'text': cell_value
                })

            # Use LLM to filter out section titles/labels, keeping only actual questions
            questions = []
            row_indices = []

            if all_rows:
                filtered_rows = self._filter_questions_with_llm(all_rows, sheet_name)
                questions = [r['text'] for r in filtered_rows]
                row_indices = [r['row_num'] for r in filtered_rows]
            
            if not questions:
                logger.warning(f"Sheet '{sheet_name}' has no questions, skipping")
                continue
            
            # Limit questions per sheet
            if len(questions) > 100:
                logger.warning(f"Sheet '{sheet_name}' has {len(questions)} questions, limiting to 100")
                questions = questions[:100]
                row_indices = row_indices[:100]
            
            # Create SheetData with column mapping and row indices
            sheet_data = SheetData(
                sheet_name=sheet_name,
                sheet_index=len(sheets),  # Reindex after filtering
                questions=questions,
                answers=[None] * len(questions),
                cell_states=[CellState.PENDING] * len(questions),
                question_col_index=question_col,
                response_col_index=response_col,
                documentation_col_index=doc_col,
                row_indices=row_indices,
                header_row_num=header_row_num
            )
            sheets.append(sheet_data)
        
        if not sheets:
            raise ExcelFormatError("No visible sheets with questions found")
        
        # Limit total sheets
        if len(sheets) > 10:
            logger.warning(f"Workbook has {len(sheets)} sheets, limiting to 10")
            sheets = sheets[:10]
        
        total_questions = sum(len(s.questions) for s in sheets)
        logger.info(f"Loaded {len(sheets)} sheets with {total_questions} total questions")
        
        return WorkbookData(file_path=file_path, sheets=sheets)

    def _find_header_row(self, ws) -> int:
        """Find the row containing column headers using LLM.

        Uses Azure AI to intelligently identify which row contains the actual column headers.

        Args:
            ws: openpyxl worksheet object

        Returns:
            Row number (1-based) of the header row
        """
        import asyncio
        import threading

        # Get first 10 rows
        rows_data = []
        for row_num in range(1, min(11, ws.max_row + 1)):
            row_values = []
            for cell in ws[row_num]:
                val = str(cell.value) if cell.value else ""
                row_values.append(val[:100])  # Limit cell length
            if any(row_values):  # Skip completely empty rows
                rows_data.append({
                    'row_number': row_num,
                    'values': row_values[:10]  # Limit to first 10 columns
                })

        if not rows_data:
            return 1

        # Build prompt for LLM
        rows_text = "\n".join([
            f"Row {r['row_number']}: {' | '.join([v for v in r['values'] if v])}"
            for r in rows_data
        ])

        prompt = f"""You are analyzing an Excel spreadsheet to find the header row.

Here are the first few rows:

{rows_text}

Which row number contains the column headers (the row with column names like "Question", "Answer", "Response", "Index", etc.)?

Respond with ONLY the row number as a single integer, nothing else."""

        try:
            # Run async code in a separate thread to avoid event loop conflicts
            async def do_header_detection():
                return await self._ask_llm_for_header_row(prompt)

            result = self._run_async_in_thread_v2(do_header_detection)

            # Parse the response
            row_num = int(result.strip())
            if 1 <= row_num <= min(10, ws.max_row):
                logger.info(f"LLM identified header row: {row_num}")
                return row_num
            else:
                logger.warning(f"LLM returned invalid row number {row_num}, using fallback")
                return self._find_header_row_fallback(ws)
        except Exception as e:
            logger.warning(f"LLM header detection failed: {e}, using fallback")
            return self._find_header_row_fallback(ws)

    async def _ask_llm_for_header_row(self, prompt: str) -> str:
        """Ask LLM to identify header row."""
        from agent_framework import ChatAgent
        from agent_framework_azure_ai import AzureAIAgentClient
        from azure.ai.projects.aio import AIProjectClient
        from azure.identity.aio import DefaultAzureCredential
        from utils.config import config_manager

        # Create a fresh Azure client for this thread's event loop
        endpoint = config_manager.get_azure_endpoint()
        model_deployment = config_manager.get_model_deployment()
        credential = DefaultAzureCredential()

        project_client = AIProjectClient(
            endpoint=endpoint,
            credential=credential
        )

        azure_client = AzureAIAgentClient(
            project_client=project_client,
            credential=credential,
            model_deployment_name=model_deployment
        )

        # Create a ChatAgent for this task
        agent = ChatAgent(
            chat_client=azure_client,
            name="Header Row Identifier",
            instructions="You are a data analysis assistant that identifies column headers in spreadsheets."
        )

        # Get the response
        response = await agent.run(prompt)

        return response.text if response else ""

    def _find_header_row_fallback(self, ws) -> int:
        """Fallback method to find header row using simple heuristics."""
        # Look for a row with multiple non-numeric values
        for row_num in range(1, min(11, ws.max_row + 1)):
            non_empty_count = 0
            for cell in ws[row_num]:
                if cell.value and not str(cell.value).isdigit():
                    non_empty_count += 1
            if non_empty_count >= 2:
                return row_num
        return 1

    def _filter_questions_with_llm(self, rows: List[dict], sheet_name: str) -> List[dict]:
        """Filter out section titles/labels, keeping only actual questions using LLM.

        Args:
            rows: List of dicts with 'row_num' and 'text' keys
            sheet_name: Name of the sheet being processed

        Returns:
            Filtered list of rows containing only actual questions
        """
        if not rows:
            return []

        # Build list of rows for LLM analysis
        rows_text = "\n".join([
            f"{i+1}. (Row {r['row_num']}): {r['text'][:200]}"
            for i, r in enumerate(rows)
        ])

        prompt = f"""You are analyzing an Excel questionnaire to identify actual questions that need answers.

Here are the rows from the "{sheet_name}" sheet Question column:

{rows_text}

Your task: Identify which rows contain items that need to be answered in a questionnaire.

EXCLUDE ONLY these types of rows:
- Pure section titles with no content expectation (e.g., "Platform Overview", "Section 1: Background")
- Instructional text for the entire section (e.g., "Answer all questions below", "Provide details in Column C")
- Decorative separators (e.g., "---", "===", "***")

INCLUDE these types of rows (these ALL need answers):
- Direct questions (e.g., "What is your platform name?", "Do you support training?")
- Imperative prompts (e.g., "Describe your capabilities", "List your differentiators", "Explain how...")
- Feature/capability items in a checklist format (e.g., "Automated ML pipeline orchestration", "Feature store support")
- Any statement that expects the user to provide information, even if not phrased as a question
- Items that would have a yes/no, description, or details as an answer

IMPORTANT RULES:
- If it's a feature, capability, or technical item that expects a response (yes/no/details), INCLUDE it
- If it's just organizing the questionnaire into sections without expecting an answer, EXCLUDE it
- When in doubt, INCLUDE it (be permissive, not restrictive)

Respond with ONLY a JSON array of the row numbers (from the original "Row X" labels) that need answers.
Example response format: [4, 5, 9, 10, 14]

If none of the rows need answers, respond with: []"""

        try:
            # Run async code in a separate thread to avoid event loop conflicts
            async def do_filtering():
                return await self._ask_llm_for_filtering(prompt)

            result = self._run_async_in_thread_v2(do_filtering)

            # Parse JSON response
            import json
            try:
                question_row_nums = json.loads(result.strip())
                if not isinstance(question_row_nums, list):
                    logger.warning(f"LLM returned invalid format, using all rows")
                    return rows

                # Filter to only rows with matching row numbers
                filtered = [r for r in rows if r['row_num'] in question_row_nums]
                logger.info(f"LLM filtered {len(rows)} rows down to {len(filtered)} questions in sheet '{sheet_name}'")
                return filtered

            except json.JSONDecodeError as e:
                logger.warning(f"Failed to parse LLM filtering response as JSON: {e}, using all rows")
                return rows

        except Exception as e:
            logger.warning(f"LLM filtering failed: {e}, using all rows")
            return rows

    async def _ask_llm_for_filtering(self, prompt: str) -> str:
        """Ask LLM to identify which rows are actual questions."""
        from agent_framework import ChatAgent
        from agent_framework_azure_ai import AzureAIAgentClient
        from azure.ai.projects.aio import AIProjectClient
        from azure.identity.aio import DefaultAzureCredential
        from utils.config import config_manager

        # Create a fresh Azure client for this thread's event loop
        endpoint = config_manager.get_azure_endpoint()
        model_deployment = config_manager.get_model_deployment()
        credential = DefaultAzureCredential()

        project_client = AIProjectClient(
            endpoint=endpoint,
            credential=credential
        )

        azure_client = AzureAIAgentClient(
            project_client=project_client,
            credential=credential,
            model_deployment_name=model_deployment
        )

        # Create a ChatAgent for this task
        agent = ChatAgent(
            chat_client=azure_client,
            name="Question Filter",
            instructions="You are a data analysis assistant that identifies actual questions in questionnaires, distinguishing them from section titles and labels."
        )

        # Get the response
        response = await agent.run(prompt)

        return response.text if response else ""

    # Constants for section header detection heuristics
    _MIN_TEXT_LENGTH = 3
    _MAX_NUMBERED_HEADER_LENGTH = 60
    _MAX_UPPERCASE_HEADER_WORDS = 6
    _MAX_TITLE_HEADER_WORDS = 5
    
    def _is_section_header(self, text: str) -> bool:
        """Determine if text appears to be a section header rather than a question.
        
        Uses heuristic rules to detect common section header patterns.
        
        Args:
            text: The cell text to analyze
            
        Returns:
            True if the text appears to be a section header, False otherwise
        """
        import re
        
        text = text.strip()
        
        # Empty or very short text
        if len(text) < self._MIN_TEXT_LENGTH:
            return True
        
        # Pattern 1: Starts with common section prefixes like "Section", "Part", "Chapter"
        section_prefixes = (
            'section ', 'part ', 'chapter ', 'category ', 'topic ', 
            'section:', 'part:', 'chapter:', 'category:', 'topic:',
            '---', '===', '***', '###'
        )
        if text.lower().startswith(section_prefixes):
            return True
        
        # Pattern 2: Entirely wrapped in decorators (dashes, equals, asterisks)
        if re.match(r'^[-=*#]+\s*.*\s*[-=*#]+$', text):
            return True
        
        # Pattern 3: Numbered sections like "1.", "1.1", "I.", "A." followed by short title
        if re.match(r'^(?:\d+\.|\d+\.\d+|[IVXLCDM]+\.|[A-Z]\.)\s+[^?]*$', text):
            # Only consider it a header if it doesn't end with '?'
            if not text.rstrip().endswith('?'):
                # And if it's relatively short without punctuation at end
                if len(text) < self._MAX_NUMBERED_HEADER_LENGTH and not text.rstrip().endswith(('.', '!', '?')):
                    return True
        
        # Pattern 4: All uppercase text (likely a header)
        if text.isupper() and len(text.split()) <= self._MAX_UPPERCASE_HEADER_WORDS:
            return True
        
        # Pattern 5: Text that contains only title-like content (no question marks, short)
        # and looks like a category label
        words = text.split()
        if (len(words) <= self._MAX_TITLE_HEADER_WORDS and 
            not text.endswith('?') and 
            text.istitle() and 
            ':' in text):
            return True
        
        return False
    
    def save_workbook(self, workbook_data: WorkbookData, output_path: str = None) -> str:
        """Save all answers back to the Excel file.
        
        Args:
            workbook_data: WorkbookData with completed answers
            output_path: Optional custom output path. If None, saves to original file.
            
        Returns:
            The path where the file was saved
            
        Raises:
            ExcelFormatError: If workbook structure changed
            IOError: If file cannot be written
        """
        # Use provided output path or default to original file path
        save_path = output_path if output_path else workbook_data.file_path
        try:
            wb = openpyxl.load_workbook(workbook_data.file_path)
        except Exception as e:
            raise ExcelFormatError(f"Cannot reopen Excel file: {e}")
        
        # Validate that sheets still exist
        existing_sheet_names = set(wb.sheetnames)
        for sheet_data in workbook_data.sheets:
            if sheet_data.sheet_name not in existing_sheet_names:
                raise ExcelFormatError(f"Sheet '{sheet_data.sheet_name}' no longer exists in file")
        
        # Write answers to identified response columns
        for sheet_data in workbook_data.sheets:
            ws = wb[sheet_data.sheet_name]
            
            # Determine response column (default to column B if not specified)
            response_col = sheet_data.response_col_index if sheet_data.response_col_index is not None else 1
            response_col_letter = openpyxl.utils.get_column_letter(response_col + 1)  # Convert 0-based to 1-based
            
            # Write header if response column is empty
            if ws.cell(row=1, column=response_col + 1).value is None:
                ws.cell(row=1, column=response_col + 1, value="Response")
            
            # Write answers to the correct rows using stored row indices
            for i, answer in enumerate(sheet_data.answers):
                if answer:
                    # Use the original row index if available, otherwise fall back to sequential
                    if sheet_data.row_indices and i < len(sheet_data.row_indices):
                        row_num = sheet_data.row_indices[i]
                    else:
                        row_num = i + 2  # Fallback to sequential starting from row 2
                    ws.cell(row=row_num, column=response_col + 1, value=answer)
        
        try:
            wb.save(save_path)
            logger.info(f"Saved workbook to {save_path}")
        except Exception as e:
            raise IOError(f"Cannot save Excel file: {e}")
        finally:
            wb.close()
        
        return save_path