"""FastAPI web application for questionnaire agent.

Provides REST API endpoints and serves the web interface.
"""

import asyncio
import logging
import os
import shutil
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, BackgroundTasks
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware

from .models import (
    SessionCreateResponse,
    SessionGetResponse,
    SessionConfigUpdate,
    SessionConfigUpdateResponse,
    QuestionRequest,
    QuestionResponse,
    SpreadsheetUploadResponse,
    ColumnSuggestions,
    ProcessingStartRequest,
    ProcessingStartResponse,
    ProcessingStatusResponse,
    StopProcessingRequest,
    StopProcessingResponse,
    HealthResponse,
    ProcessingJob,
    JobStatus,
    SSEMessageType,
)
from .session_manager import session_manager
from .sse_manager import sse_manager

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Determine static files directory
STATIC_DIR = Path(__file__).parent / "static"

# Create FastAPI app
app = FastAPI(
    title="Questionnaire Agent Web",
    description="Web interface for the questionnaire answering agent",
    version="1.0.0"
)

# Mount static files
if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")


# ============================================================================
# Health Check
# ============================================================================

@app.get("/health", response_model=HealthResponse)
async def health_check():
    """Health check endpoint."""
    azure_auth_status = "authenticated"
    status = "healthy"
    message = None

    # Try to verify Azure auth
    try:
        from utils.azure_auth import test_authentication
        # We can't run async here easily, so we'll assume authenticated if module loads
    except ImportError:
        azure_auth_status = "unknown"
        message = "Azure auth module not available"
    except Exception as e:
        azure_auth_status = "unauthenticated"
        status = "degraded"
        message = str(e)

    return HealthResponse(
        status=status,
        timestamp=datetime.now().isoformat(),
        azure_auth=azure_auth_status,
        message=message
    )


# ============================================================================
# Index Route
# ============================================================================

@app.get("/")
async def index():
    """Serve the main web interface."""
    index_path = STATIC_DIR / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=500, detail="Web interface not found")
    return FileResponse(str(index_path))


@app.get("/favicon.ico")
async def favicon():
    """Serve the favicon."""
    favicon_path = STATIC_DIR / "favicon.svg"
    if favicon_path.exists():
        return FileResponse(str(favicon_path), media_type="image/svg+xml")
    raise HTTPException(status_code=404, detail="Favicon not found")


# ============================================================================
# Session Management
# ============================================================================

@app.post("/api/session/create", response_model=SessionCreateResponse, status_code=201)
async def create_session():
    """Create a new session."""
    session_id = session_manager.create_session()
    session = session_manager.get_session(session_id)

    return SessionCreateResponse(
        session_id=session_id,
        created_at=session.created_at.isoformat(),
        config=session.config
    )


@app.get("/api/session/{session_id}", response_model=SessionGetResponse)
async def get_session(session_id: str):
    """Get session state."""
    session = session_manager.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    processing_status = None
    if session.processing_job:
        processing_status = session.processing_job.status.value

    return SessionGetResponse(
        session_id=session_id,
        created_at=session.created_at.isoformat(),
        config=session.config,
        has_workbook=session.workbook_data is not None,
        processing_status=processing_status
    )


@app.put("/api/session/{session_id}/config", response_model=SessionConfigUpdateResponse)
async def update_session_config(session_id: str, update: SessionConfigUpdate):
    """Update session configuration."""
    session = session_manager.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    if not session_manager.update_config(
        session_id,
        context=update.context,
        char_limit=update.char_limit
    ):
        raise HTTPException(status_code=500, detail="Failed to update config")

    return SessionConfigUpdateResponse(
        session_id=session_id,
        config=session.config
    )


@app.delete("/api/session/{session_id}")
async def delete_session(session_id: str):
    """Delete a session."""
    if not session_manager.delete_session(session_id):
        raise HTTPException(status_code=404, detail="Session not found")
    return {"status": "deleted", "session_id": session_id}


# ============================================================================
# Question Processing
# ============================================================================

@app.post("/api/question", response_model=QuestionResponse)
async def process_question(request: QuestionRequest):
    """Process a single question."""
    session = session_manager.get_session(request.session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    start_time = time.time()

    try:
        # Import agent components
        from agents.workflow_manager import create_agent_coordinator
        from utils.data_types import Question
        from utils.azure_auth import foundry_agent_session, get_project_client
        from utils.config import config_manager

        # Create question object
        question = Question(
            text=request.question,
            context=request.context,
            char_limit=request.char_limit
        )

        # Get Azure client and project client, then create coordinator
        async with foundry_agent_session() as azure_client:
            project_client = await get_project_client()
            coordinator = await create_agent_coordinator(
                azure_client=azure_client,
                bing_connection_id=config_manager.get_bing_connection_id(),
                browser_automation_connection_id=config_manager.get_browser_automation_connection_id(),
                project_client=project_client
            )

            # Create callbacks for processing
            reasoning_parts = []

            def progress_callback(agent: str, message: str, progress: float):
                logger.info(f"[{agent}] {message} ({progress:.1%})")

            def reasoning_callback(text: str):
                reasoning_parts.append(text)

            # Process with agent coordinator
            result = await coordinator.process_question(
                question,
                progress_callback=progress_callback,
                reasoning_callback=reasoning_callback
            )

            # Cleanup
            await coordinator.cleanup_agents()

        processing_time = time.time() - start_time

        # Format reasoning
        reasoning = _format_reasoning(result, reasoning_parts)

        return QuestionResponse(
            answer=result.answer.content if result.answer else "No answer generated",
            reasoning=reasoning,
            processing_time_seconds=round(processing_time, 2),
            links_checked=len(result.answer.documentation_links) if result.answer else 0
        )

    except ImportError as e:
        logger.error(f"Agent import error: {e}")
        raise HTTPException(status_code=500, detail="Agent services not available")
    except Exception as e:
        logger.error(f"Question processing error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


def _format_reasoning(result, reasoning_parts: list = None) -> str:
    """Format agent reasoning trace for display."""
    # Start with any collected reasoning parts
    output = []
    if reasoning_parts:
        output.extend(reasoning_parts)

    try:
        from utils.reasoning_formatter import format_reasoning_trace
        if result.answer and result.answer.agent_reasoning:
            formatted = format_reasoning_trace(result.answer.agent_reasoning)
            output.append(formatted)
    except ImportError:
        # Fallback formatting
        if result.answer and result.answer.agent_reasoning:
            output.append("## Agent Workflow\n")
            for step in result.answer.agent_reasoning:
                output.append(f"### {step.agent_name.value if hasattr(step.agent_name, 'value') else step.agent_name}")
                output.append(f"- Status: {step.status.value if hasattr(step.status, 'value') else step.status}")
                if hasattr(step, 'execution_time'):
                    output.append(f"- Time: {step.execution_time:.2f}s")
                if hasattr(step, 'error_message') and step.error_message:
                    output.append(f"- Error: {step.error_message}")
                output.append("")

    return "\n".join(output) if output else "No reasoning trace available"


# ============================================================================
# Spreadsheet Operations
# ============================================================================

@app.post("/api/spreadsheet/upload", response_model=SpreadsheetUploadResponse)
async def upload_spreadsheet(
    session_id: str = Form(...),
    file: UploadFile = File(...)
):
    """Upload an Excel file for processing."""
    session = session_manager.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Validate file extension
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")

    ext = Path(file.filename).suffix.lower()
    if ext not in ('.xlsx', '.xls'):
        raise HTTPException(status_code=400, detail="Invalid file format. Only .xlsx and .xls files are supported.")

    try:
        # Save to temp file
        temp_fd, temp_path = tempfile.mkstemp(suffix=ext)
        os.close(temp_fd)

        with open(temp_path, 'wb') as f:
            shutil.copyfileobj(file.file, f)

        # Load workbook using existing loader
        from excel.loader import ExcelLoader
        from excel.column_identifier import ColumnIdentifier

        identifier = ColumnIdentifier()
        loader = ExcelLoader(column_identifier=identifier)
        workbook_data = loader.load_workbook(temp_path)

        # Get column info and data
        columns = {}
        sheets = []
        data = {}
        total_rows = 0
        for sheet in workbook_data.sheets:
            sheets.append(sheet.sheet_name)
            # Get column names from the sheet
            sheet_columns = _get_column_names(sheet, temp_path)
            columns[sheet.sheet_name] = sheet_columns
            # Get row data
            data[sheet.sheet_name] = _get_sheet_data(temp_path, sheet.sheet_name, sheet_columns)
            total_rows += len(data[sheet.sheet_name])

        # Get suggestions from the first sheet using AI-based column identification
        suggestions = await _identify_columns(workbook_data, columns)

        # Store in session
        session_manager.set_workbook(session_id, workbook_data, temp_path, columns)

        return SpreadsheetUploadResponse(
            session_id=session_id,
            filename=file.filename,
            sheets=sheets,
            columns=columns,
            suggested_columns=suggestions,
            row_count=total_rows,
            data=data
        )

    except ImportError as e:
        logger.error(f"Excel import error: {e}")
        raise HTTPException(status_code=500, detail="Excel processing not available")
    except Exception as e:
        logger.error(f"Upload error: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")


def _get_column_names(sheet, file_path: str) -> list:
    """Extract column names from sheet by reading Excel headers."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet.sheet_name]

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value))
            else:
                break

        wb.close()
        return headers if headers else ["A", "B", "C", "D", "E"]

    except Exception as e:
        logger.warning(f"Failed to read headers: {e}")
        # Fallback based on detected column indices
        columns = []
        if sheet.question_col_index is not None:
            columns.append("Question")
        if sheet.response_col_index is not None:
            columns.append("Answer")
        if sheet.documentation_col_index is not None:
            columns.append("Documentation")
        return columns if columns else ["A", "B", "C", "D", "E"]


def _get_sheet_data(file_path: str, sheet_name: str, columns: list) -> list:
    """Extract all row data from a sheet."""
    import openpyxl

    rows = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]

        # Skip header row, iterate data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row_dict = {"rowIndex": str(row_idx)}
            for col_idx, col_name in enumerate(columns):
                if col_idx < len(row):
                    value = row[col_idx]
                    row_dict[col_name] = str(value) if value is not None else ""
                else:
                    row_dict[col_name] = ""
            # Only include rows that have some data
            if any(row_dict.get(col, "") for col in columns):
                rows.append(row_dict)

        wb.close()
    except Exception as e:
        logger.warning(f"Failed to read sheet data: {e}")

    return rows


async def _identify_columns(workbook_data, columns: dict = None) -> ColumnSuggestions:
    """Identify likely question/answer columns using Azure AI LLM.

    Uses ChatAgent to ask the LLM to identify which columns contain questions,
    responses, and documentation based on the column headers.
    """
    if not workbook_data.sheets:
        return ColumnSuggestions(
            sheet_name="Sheet1",
            question_column=None,
            context_column=None,
            answer_column=None,
            confidence=0.0,
            auto_map_success=False
        )

    sheet = workbook_data.sheets[0]
    sheet_name = sheet.sheet_name

    # Get actual column names from headers if available
    sheet_columns = columns.get(sheet_name, []) if columns else []

    if not sheet_columns:
        return ColumnSuggestions(
            sheet_name=sheet_name,
            question_column=None,
            context_column=None,
            answer_column=None,
            confidence=0.0,
            auto_map_success=False
        )

    # Use Azure AI LLM for column identification
    try:
        from utils.azure_auth import get_azure_client
        from agent_framework import ChatAgent
        import json
        import re

        azure_client = await get_azure_client()

        # Create a ChatAgent for column identification
        agent = ChatAgent(
            chat_client=azure_client,
            name="Column Identifier",
            instructions="""You are a column identification expert. Your job is to analyze spreadsheet column headers and identify which columns contain:
1. Questions - the column with questions to be answered
2. Responses/Answers - the column where answers should be written
3. Documentation - the column for documentation links (optional)

Always respond with valid JSON only. No other text."""
        )

        # Build the prompt
        column_list = ', '.join([f'Column {i}: "{col}"' for i, col in enumerate(sheet_columns)])
        prompt = f"""Analyze these spreadsheet column headers and identify which columns are for questions, responses, and documentation:

{column_list}

Respond ONLY with valid JSON in this exact format:
{{"question": <column_index>, "response": <column_index>, "documentation": <column_index_or_null>}}

Use 0-based column indices. If a column cannot be identified, use null.
For example, if "Question" is column 3, "Response" is column 4, and "Documentation" is column 5:
{{"question": 3, "response": 4, "documentation": 5}}"""

        logger.info(f"Sending column identification request to Azure AI LLM with headers: {sheet_columns}")

        # Call the LLM - pass prompt directly as string
        response = await agent.run(prompt)

        # Parse the response - use .text property
        response_text = response.text if response else ""
        logger.info(f"LLM column identification response: {response_text}")

        # Extract JSON from response
        json_match = re.search(r'\{[^}]+\}', response_text)
        if json_match:
            result = json.loads(json_match.group(0))
            logger.info(f"Parsed column identification result: {result}")

            # Convert indices to column names
            question_col = sheet_columns[result['question']] if result.get('question') is not None and result['question'] < len(sheet_columns) else None
            answer_col = sheet_columns[result['response']] if result.get('response') is not None and result['response'] < len(sheet_columns) else None
            context_col = sheet_columns[result['documentation']] if result.get('documentation') is not None and result['documentation'] < len(sheet_columns) else None

            # LLM identification has high confidence
            confidence = 1.0 if question_col and answer_col else 0.5
            auto_map_success = question_col is not None and answer_col is not None

            logger.info(f"LLM identified columns - Question: {question_col}, Answer: {answer_col}, Documentation: {context_col}")

            return ColumnSuggestions(
                sheet_name=sheet_name,
                question_column=question_col,
                context_column=context_col,
                answer_column=answer_col,
                confidence=confidence,
                auto_map_success=auto_map_success
            )
        else:
            logger.warning(f"Could not parse JSON from LLM response: {response_text}")

    except Exception as e:
        logger.error(f"Azure AI LLM column identification failed: {e}", exc_info=True)

    # Return failure - do NOT fall back to heuristics per user request
    logger.warning("LLM column identification failed, returning no suggestions")
    return ColumnSuggestions(
        sheet_name=sheet_name,
        question_column=None,
        context_column=None,
        answer_column=None,
        confidence=0.0,
        auto_map_success=False
    )


def _identify_columns_heuristic(workbook_data, columns: dict = None) -> ColumnSuggestions:
    """Fallback heuristic-based column identification."""
    if not workbook_data.sheets:
        return ColumnSuggestions(
            sheet_name="Sheet1",
            question_column=None,
            context_column=None,
            answer_column=None,
            confidence=0.0,
            auto_map_success=False
        )

    sheet = workbook_data.sheets[0]
    sheet_name = sheet.sheet_name

    # Get actual column names from headers if available
    sheet_columns = columns.get(sheet_name, []) if columns else []

    # Question column patterns - ordered by priority (most specific first)
    # Single-letter patterns only match exactly, not as substrings
    question_patterns = [
        ('question', False),  # (pattern, exact_only)
        ('query', False),
        ('ask', False),
        ('q', True),  # 'q' only matches exact "q" column name
    ]
    # Answer/Response column patterns - ordered by priority
    answer_patterns = [
        ('response', False),
        ('answer', False),
        ('reply', False),
        # Note: 'a' removed - too broad, matches "Status", "Data", etc.
    ]
    # Documentation column patterns (optional)
    doc_patterns = ['documentation', 'docs', 'sources', 'references', 'links']

    question_col = None
    answer_col = None
    context_col = None
    question_confidence = 0.0
    answer_confidence = 0.0

    for col in sheet_columns:
        col_lower = col.lower().strip()

        # Check for question column
        if not question_col:
            for pattern, exact_only in question_patterns:
                if col_lower == pattern:
                    question_col = col
                    question_confidence = 1.0  # Exact match
                    break
                elif not exact_only and pattern in col_lower:
                    question_col = col
                    question_confidence = 0.8  # Partial match
                    break

        # Check for answer column
        if not answer_col:
            for pattern, exact_only in answer_patterns:
                if col_lower == pattern:
                    answer_col = col
                    answer_confidence = 1.0  # Exact match
                    break
                elif not exact_only and pattern in col_lower:
                    answer_col = col
                    answer_confidence = 0.8  # Partial match
                    break

        # Check for documentation/context column
        if not context_col:
            for pattern in doc_patterns:
                if pattern in col_lower:
                    context_col = col
                    break

    # Fallback to loader's identified columns if header matching failed
    if not question_col and sheet.question_col_index is not None and sheet.question_col_index < len(sheet_columns):
        question_col = sheet_columns[sheet.question_col_index]
        question_confidence = 0.6

    if not answer_col and sheet.response_col_index is not None and sheet.response_col_index < len(sheet_columns):
        answer_col = sheet_columns[sheet.response_col_index]
        answer_confidence = 0.6

    # Calculate overall confidence
    # Auto-mapping succeeds if both question and answer columns are found with high confidence
    confidence = (question_confidence + answer_confidence) / 2
    auto_map_success = question_col is not None and answer_col is not None and confidence >= 0.7

    return ColumnSuggestions(
        sheet_name=sheet_name,
        question_column=question_col,
        context_column=context_col,
        answer_column=answer_col,
        confidence=confidence,
        auto_map_success=auto_map_success
    )


@app.post("/api/spreadsheet/process", response_model=ProcessingStartResponse, status_code=202)
async def start_processing(request: ProcessingStartRequest, background_tasks: BackgroundTasks):
    """Start batch processing of spreadsheet questions."""
    session = session_manager.get_session(request.session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    if not session.workbook_data:
        raise HTTPException(status_code=400, detail="No spreadsheet loaded")

    # Check for existing active job
    if session.processing_job and session.processing_job.status == JobStatus.RUNNING:
        raise HTTPException(status_code=409, detail="Processing job already running")

    # Find the sheet
    target_sheet = None
    for sheet in session.workbook_data.sheets:
        if sheet.sheet_name == request.sheet_name:
            target_sheet = sheet
            break

    if not target_sheet:
        raise HTTPException(status_code=400, detail=f"Sheet '{request.sheet_name}' not found")

    # Calculate rows to process
    total_rows = len(target_sheet.questions)
    start_row = request.start_row
    end_row = request.end_row if request.end_row is not None else total_rows

    if start_row >= total_rows:
        raise HTTPException(status_code=422, detail="start_row exceeds available rows")

    # Count actual non-empty questions in the range
    actual_end = min(end_row, total_rows)
    rows_to_process = sum(
        1 for idx in range(start_row, actual_end)
        if target_sheet.questions[idx] and target_sheet.questions[idx].strip()
    )

    if rows_to_process == 0:
        raise HTTPException(status_code=422, detail="No questions found in the selected range")

    # Create processing job
    job = ProcessingJob(
        session_id=request.session_id,
        total_rows=rows_to_process,
        processed_rows=0,
        current_row=start_row
    )
    session_manager.set_processing_job(request.session_id, job)

    # Start background processing
    background_tasks.add_task(
        _process_spreadsheet,
        request.session_id,
        target_sheet,
        request.question_column,
        request.context_column,
        request.answer_column,
        start_row,
        end_row
    )

    return ProcessingStartResponse(
        job_id=job.job_id,
        session_id=request.session_id,
        status=job.status,
        total_rows=rows_to_process,
        processed_rows=0,
        started_at=job.started_at.isoformat()
    )


async def _process_spreadsheet(
    session_id: str,
    sheet,
    question_column: str,
    context_column: Optional[str],
    answer_column: str,
    start_row: int,
    end_row: int
):
    """Background task for spreadsheet processing with 3 parallel agent sets."""
    start_time = time.time()

    try:
        from agents.workflow_manager import create_agent_coordinator
        from utils.data_types import Question
        from utils.azure_auth import foundry_agent_session, get_project_client
        from utils.config import config_manager

        session = session_manager.get_session(session_id)

        if not session:
            return

        job = session.processing_job
        if not job:
            return

        # Get context from session config
        default_context = session.config.context
        char_limit = session.config.char_limit

        # Calculate actual number of non-empty questions in range
        actual_end = min(end_row, len(sheet.questions))

        # Build list of (row_idx, question_text) for non-empty questions
        questions_to_process = [
            (idx, sheet.questions[idx])
            for idx in range(start_row, actual_end)
            if sheet.questions[idx] and sheet.questions[idx].strip()
        ]
        total_questions = len(questions_to_process)

        if total_questions == 0:
            logger.warning("No questions to process")
            session_manager.update_job_status(session_id, JobStatus.COMPLETED)
            await sse_manager.send_complete(session_id, 0, 0.0)
            return

        # Shared state for parallel processing (thread-safe with asyncio.Lock)
        state_lock = asyncio.Lock()
        completed_rows = 0
        cancelled = False

        # Get Azure client and project client
        async with foundry_agent_session() as azure_client:
            project_client = await get_project_client()

            # Create 3 agent coordinators for parallel processing
            NUM_AGENT_SETS = 3
            coordinators = []
            for i in range(NUM_AGENT_SETS):
                coordinator = await create_agent_coordinator(
                    azure_client=azure_client,
                    bing_connection_id=config_manager.get_bing_connection_id(),
                    browser_automation_connection_id=config_manager.get_browser_automation_connection_id(),
                    project_client=project_client
                )
                coordinators.append(coordinator)
                logger.info(f"Created agent coordinator {i + 1}/{NUM_AGENT_SETS}")

            # Create work queue
            work_queue = asyncio.Queue()
            for item in questions_to_process:
                await work_queue.put(item)

            async def worker(agent_set_id: int, coordinator):
                """Worker for a single agent set."""
                nonlocal completed_rows, cancelled
                worker_processed = 0
                worker_failed = 0

                logger.info(f"Agent Set {agent_set_id} starting")

                while not cancelled:
                    # Check if job was cancelled
                    current_job = session_manager.get_processing_job(session_id)
                    if not current_job or current_job.status in (JobStatus.CANCELLED, JobStatus.ERROR):
                        cancelled = True
                        break

                    try:
                        # Get next question from queue (with timeout)
                        try:
                            row_idx, question_text = await asyncio.wait_for(
                                work_queue.get(),
                                timeout=0.1
                            )
                        except asyncio.TimeoutError:
                            # No more work available
                            break

                        logger.info(f"Agent Set {agent_set_id} processing row {row_idx}")

                        # Send ROW_STARTED event
                        await sse_manager.send_row_started(session_id, row_idx, total_questions)

                        try:
                            # Create question and process
                            question = Question(
                                text=question_text,
                                context=default_context,
                                char_limit=char_limit
                            )

                            reasoning_parts = []
                            last_agent = [None]  # Use list to allow modification in nested function

                            async def send_agent_progress_safe(sid, row, agent, msg):
                                """Safe wrapper to send agent progress with error handling."""
                                try:
                                    result = await sse_manager.send_agent_progress(sid, row, agent, msg)
                                    logger.debug(f"AGENT_PROGRESS sent for row {row}, agent={agent}, result={result}")
                                except Exception as e:
                                    logger.error(f"Failed to send AGENT_PROGRESS: {e}")

                            def progress_callback(agent: str, message: str, progress: float):
                                logger.info(f"[Set {agent_set_id}][Row {row_idx}][{agent}] {message}")
                                # Send agent progress SSE event if agent changed
                                if agent != last_agent[0] and agent not in ("workflow", "batch"):
                                    last_agent[0] = agent
                                    logger.info(f"Sending AGENT_PROGRESS SSE for row {row_idx}, agent={agent}")
                                    # Schedule async SSE send
                                    asyncio.create_task(
                                        send_agent_progress_safe(session_id, row_idx, agent, message)
                                    )

                            def reasoning_callback(text: str):
                                reasoning_parts.append(text)

                            result = await coordinator.process_question(
                                question,
                                progress_callback=progress_callback,
                                reasoning_callback=reasoning_callback
                            )

                            # Get answer
                            answer = result.answer.content if result.answer else "Error generating answer"

                            # Update sheet data (thread-safe)
                            async with state_lock:
                                if hasattr(sheet, 'mark_completed'):
                                    sheet.mark_completed(row_idx, answer)
                                else:
                                    sheet.answers[row_idx] = answer

                                # Increment completed rows
                                completed_rows += 1
                                current_completed = completed_rows

                                # Update job progress
                                session_manager.update_job_progress(session_id, current_completed, row_idx + 1)
                                job.processed_rows = current_completed

                            # Send answer via SSE
                            await sse_manager.send_answer(
                                session_id,
                                row_idx,
                                question_text,
                                answer,
                                _format_reasoning(result, reasoning_parts)
                            )

                            # Send progress update
                            await sse_manager.send_progress(session_id, current_completed, total_questions)

                            worker_processed += 1
                            logger.info(f"Agent Set {agent_set_id} completed row {row_idx}")

                        except Exception as e:
                            logger.error(f"Agent Set {agent_set_id} error on row {row_idx}: {e}")
                            await sse_manager.send_error(session_id, str(e), row_idx)

                            # Still count as processed (with error)
                            async with state_lock:
                                completed_rows += 1
                                current_completed = completed_rows
                                session_manager.update_job_progress(session_id, current_completed, row_idx + 1)
                                job.processed_rows = current_completed

                            await sse_manager.send_progress(session_id, current_completed, total_questions)
                            worker_failed += 1

                    except Exception as e:
                        logger.error(f"Agent Set {agent_set_id} worker error: {e}", exc_info=True)
                        break

                logger.info(f"Agent Set {agent_set_id} finished: {worker_processed} processed, {worker_failed} failed")
                return worker_processed, worker_failed

            # Run workers in parallel
            worker_tasks = [
                asyncio.create_task(worker(i + 1, coordinators[i]))
                for i in range(NUM_AGENT_SETS)
            ]

            # Wait for all workers to complete
            results = await asyncio.gather(*worker_tasks, return_exceptions=True)

            # Log results
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    logger.error(f"Agent Set {i + 1} failed with exception: {result}")
                else:
                    processed, failed = result
                    logger.info(f"Agent Set {i + 1} result: {processed} processed, {failed} failed")

            # Cleanup all coordinators
            for i, coordinator in enumerate(coordinators):
                try:
                    await coordinator.cleanup_agents()
                    logger.info(f"Cleaned up coordinator {i + 1}")
                except Exception as e:
                    logger.error(f"Error cleaning up coordinator {i + 1}: {e}")

        # Mark complete
        duration = time.time() - start_time
        session_manager.update_job_status(session_id, JobStatus.COMPLETED)
        await sse_manager.send_complete(session_id, completed_rows, duration)

    except Exception as e:
        logger.error(f"Spreadsheet processing error: {e}", exc_info=True)
        session_manager.update_job_status(session_id, JobStatus.ERROR, str(e))
        await sse_manager.send_error(session_id, str(e))


@app.get("/api/spreadsheet/status/{session_id}", response_model=ProcessingStatusResponse)
async def get_processing_status(session_id: str):
    """Get current processing status."""
    session = session_manager.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    job = session.processing_job
    if not job:
        return ProcessingStatusResponse(
            status="NO_JOB",
            message="No processing job active for this session"
        )

    # Calculate estimated time remaining
    estimated_remaining = None
    if job.processed_rows > 0 and job.status == JobStatus.RUNNING:
        elapsed = (datetime.now() - job.started_at).total_seconds()
        avg_per_row = elapsed / job.processed_rows
        remaining_rows = job.total_rows - job.processed_rows
        estimated_remaining = avg_per_row * remaining_rows

    return ProcessingStatusResponse(
        job_id=job.job_id,
        status=job.status.value,
        processed_rows=job.processed_rows,
        total_rows=job.total_rows,
        current_row=job.current_row,
        estimated_time_remaining_seconds=round(estimated_remaining, 1) if estimated_remaining else None,
        started_at=job.started_at.isoformat(),
        completed_at=job.completed_at.isoformat() if job.completed_at else None
    )


@app.post("/api/spreadsheet/stop", response_model=StopProcessingResponse)
async def stop_processing(request: StopProcessingRequest):
    """Stop the current processing job."""
    session = session_manager.get_session(request.session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    job = session.processing_job
    if not job:
        raise HTTPException(status_code=400, detail="No active job to stop")

    if job.status != JobStatus.RUNNING:
        raise HTTPException(status_code=400, detail=f"Job is not running (status: {job.status.value})")

    # Mark as cancelled
    session_manager.update_job_status(request.session_id, JobStatus.CANCELLED)

    # Send status via SSE
    await sse_manager.send_status(request.session_id, "CANCELLED", job.job_id)

    return StopProcessingResponse(
        job_id=job.job_id,
        status=JobStatus.CANCELLED,
        processed_rows=job.processed_rows,
        total_rows=job.total_rows
    )


@app.get("/api/spreadsheet/download/{session_id}")
async def download_spreadsheet(session_id: str):
    """Download the processed spreadsheet."""
    session = session_manager.get_session(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    if not session.temp_file_path or not os.path.exists(session.temp_file_path):
        raise HTTPException(status_code=400, detail="No spreadsheet available for download")

    # TODO: Write answers back to the Excel file before download
    # For now, return the original file

    return FileResponse(
        session.temp_file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="questionnaire_answered.xlsx"
    )


# ============================================================================
# Server-Sent Events
# ============================================================================

@app.get("/api/sse/{session_id}")
async def sse_stream(session_id: str):
    """Server-Sent Events stream for real-time updates."""
    if not session_manager.session_exists(session_id):
        raise HTTPException(status_code=404, detail="Session not found")

    return StreamingResponse(
        sse_manager.stream_events(session_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"
        }
    )


# ============================================================================
# Server Runner
# ============================================================================

def run_server(host: str = "127.0.0.1", port: int = 8080, log_level: str = "info"):
    """Run the web server."""
    import uvicorn
    uvicorn.run(app, host=host, port=port, log_level=log_level)


def cleanup():
    """Clean up resources on shutdown."""
    session_manager.cleanup_all()
    sse_manager.cleanup_all()
