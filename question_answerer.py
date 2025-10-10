#!/usr/bin/env python3
"""
Questionnaire Multiagent - UI Application

A windowed application that orchestrates three Azure AI Foundry agents to answer questions
with fact-checking and link validation. Supports both individual questions and Excel import/export.
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import os
import tempfile
import asyncio
import logging
import argparse
import sys
import atexit
import signal
from typing import Tuple, Optional, List, Dict, Any
from pathlib import Path
from azure.monitor.opentelemetry import configure_azure_monitor
from opentelemetry import trace
from opentelemetry.trace import Tracer

import pandas as pd
import warnings
from azure.ai.projects import AIProjectClient
from azure.identity import DefaultAzureCredential
from azure.ai.agents.models import BingGroundingTool
from dotenv import load_dotenv

# Suppress openpyxl data validation warnings - these are not actionable by users
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader', message='.*Data Validation extension.*')

# Import the resource manager for agent cleanup
from utils.resource_manager import FoundryAgentSession

# Load environment variables
load_dotenv(override=True)

class QuestionnaireAgentUI:
    """Main UI application for the Questionnaire Agent."""
    
    def __init__(self, headless_mode=False, max_retries=10, mock_mode=False):
        # Setup logging first - only for our app, not Azure SDK noise
        logging.basicConfig(level=logging.WARNING)  # Set root to WARNING to silence Azure SDK
        self.logger = logging.getLogger(__name__)
        self.logger.setLevel(logging.INFO)  # But keep our app logger at INFO
        
        # Silence specific noisy Azure loggers
        logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.WARNING)
        logging.getLogger("azure.monitor.opentelemetry.exporter").setLevel(logging.WARNING)
        logging.getLogger("azure.monitor.opentelemetry").setLevel(logging.INFO)
        
        # Store headless mode flag
        self.headless_mode = headless_mode
        
        # Store maximum retries configuration
        self.max_retries = max_retries
        
        # Store mock mode flag
        self.mock_mode = mock_mode
        
        if not headless_mode:
            # Enable high DPI awareness on Windows
            try:
                from ctypes import windll
                windll.shcore.SetProcessDpiAwareness(1)
            except:
                pass  # Ignore errors on non-Windows platforms or if not available
            
            self.root = tk.Tk()
            self.root.title("Questionnaire Multiagent")
            self.root.geometry("1200x800")
            
            # Maximize window based on OS
            import platform
            try:
                if platform.system() == "Windows":
                    self.root.state('zoomed')  # Windows maximized
                elif platform.system() == "Linux":
                    self.root.attributes('-zoomed', True)  # Linux maximized
                elif platform.system() == "Darwin":  # macOS
                    self.root.attributes('-zoomed', True)  # macOS maximized
            except tk.TclError:
                # Fallback if maximizing fails - just use the geometry
                pass
        else:
            self.root = None
        
        # OpenTelemetry tracer for Azure AI Foundry tracing
        self.tracer: Optional[Tracer] = None
        
        # Initialize tracing BEFORE Azure client initialization (skip in mock mode)
        if not self.mock_mode:
            self.initialize_tracing()
        else:
            self.tracer = None
        
        # Initialize Azure AI Project Client (skip in mock mode)
        self.project_client = None
        if not self.mock_mode:
            self.init_azure_client()
        else:
            self.logger.info("Mock mode enabled - skipping Azure client initialization")
        
        # Agent IDs will be set when agents are created
        self.question_answerer_id = None
        self.answer_checker_id = None
        self.link_checker_id = None
        
        # Agent sessions for proper resource cleanup
        self.question_answerer_session = None
        self.answer_checker_session = None
        self.link_checker_session = None
        
        # CLI output buffer for capturing agent responses
        self.cli_output = []
        
        # Setup UI only if not in headless mode
        if not headless_mode:
            self.setup_ui()
            
            # Status bar variables (only needed for GUI mode and after root is created)
            self.status_working = tk.StringVar(value="Idle")
            self.status_agent = tk.StringVar(value="")
            self.status_time = tk.StringVar(value="00:00")
            self.status_excel_input = tk.StringVar(value="")
            self.status_excel_output = tk.StringVar(value="")
            self.status_excel_question = tk.StringVar(value="")
            
            # Timer variables
            self.start_time = None
            self.timer_job = None
            
            # Now setup the status bar after StringVars are created
            self.setup_status_bar()
        else:
            # For headless mode, create dummy status variables
            self.status_working = None
            self.status_agent = None
            self.status_time = None
            self.status_excel_input = None
            self.status_excel_output = None
            self.status_excel_question = None
            self.start_time = None
            self.timer_job = None
        
        # Register cleanup handlers for proper agent resource management
        self._setup_cleanup_handlers()
        
    def _setup_cleanup_handlers(self):
        """Setup cleanup handlers for agent resources."""
        # Register cleanup on normal program exit
        atexit.register(self.cleanup_agents)
        
        # Register signal handlers for graceful shutdown
        signal.signal(signal.SIGINT, self._signal_handler)
        signal.signal(signal.SIGTERM, self._signal_handler)
        
        # Setup GUI window close protocol if in GUI mode
        if not self.headless_mode and self.root:
            self.root.protocol("WM_DELETE_WINDOW", self._on_window_close)
    
    def _signal_handler(self, signum, frame):
        """Handle shutdown signals gracefully."""
        self.logger.info(f"Received signal {signum}, cleaning up agents...")
        self.cleanup_agents()
        sys.exit(0)
    
    def _on_window_close(self):
        """Handle GUI window close event."""
        self.logger.info("Application window closing, cleaning up agents...")
        self.cleanup_agents()
        if self.root:
            self.root.destroy()
    
    def __del__(self):
        """Destructor to ensure cleanup on object destruction."""
        self.cleanup_agents()
        
    def init_azure_client(self):
        """Initialize Azure AI Project Client with credentials from .env file."""
        try:
            endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
            if not endpoint:
                raise ValueError("AZURE_OPENAI_ENDPOINT not found in environment variables")
            
            self.logger.info(f"Connecting to Azure AI Foundry endpoint: {endpoint}")
            
            credential = DefaultAzureCredential()
            self.project_client = AIProjectClient(
                endpoint=endpoint,
                credential=credential
            )
            self.logger.info("Azure AI Project Client initialized successfully")
            
        except Exception as e:
            self.logger.error(f"Failed to initialize Azure client: {e}")
            if not self.headless_mode:
                messagebox.showerror("Azure Connection Error", 
                                   f"Failed to connect to Azure AI Foundry:\n{e}\n\nPlease check your .env file and Azure credentials.")
            else:
                print(f"Error: Failed to connect to Azure AI Foundry: {e}")
                print("Please check your .env file and Azure credentials.")
                sys.exit(1)
    
    def initialize_tracing(self):
        """Initialize Azure AI Foundry tracing with Application Insights"""
        try:
            # Configure content recording based on environment variable or default to true for debugging
            content_recording = os.environ.get("AZURE_TRACING_GEN_AI_CONTENT_RECORDING_ENABLED", "true").lower()
            os.environ["AZURE_TRACING_GEN_AI_CONTENT_RECORDING_ENABLED"] = content_recording
            
            # Set Azure SDK tracing implementation to OpenTelemetry BEFORE any Azure operations
            os.environ["AZURE_SDK_TRACING_IMPLEMENTATION"] = "opentelemetry"
            
            # Set service name for Application Analytics in Azure AI Foundry
            # This maps to cloud_RoleName in Application Insights and enables the Application Analytics dashboard
            if not os.environ.get("OTEL_SERVICE_NAME"):
                os.environ["OTEL_SERVICE_NAME"] = "Questionnaire Agent V2"
            
            # Enable Azure SDK tracing with OpenTelemetry BEFORE configuring Azure Monitor
            from azure.core.settings import settings
            settings.tracing_implementation = "opentelemetry"
            
            # Get Application Insights connection string from environment variable
            connection_string = os.environ.get("APPLICATIONINSIGHTS_CONNECTION_STRING")
            
            if not connection_string:
                self.logger.warning("⚠️ APPLICATIONINSIGHTS_CONNECTION_STRING environment variable not set.")
                self.logger.warning("   Set this in your .env file - get it from Azure Portal > Application Insights > Overview")
                return False
            
            # Configure Azure Monitor tracing with enhanced settings for AI Foundry
            configure_azure_monitor(
                connection_string=connection_string,
                # Enable additional instrumentation for better metrics capture
                enable_live_metrics=True,
                # Set sampling rate to 100% to capture all operations for now
                sampling_ratio=1.0
            )
            
            # Optional: Enable telemetry to console for debugging (uncomment if needed)
            # try:
            #     from azure.ai.agents.telemetry import enable_telemetry
            #     import sys
            #     enable_telemetry(destination=sys.stdout)
            #     self.logger.info("✅ Console telemetry enabled for debugging")
            # except ImportError:
            #     pass  # azure-ai-agents telemetry not available
            
            # Create a tracer for custom spans
            self.tracer = trace.get_tracer(__name__)
            
            content_status = "enabled" if content_recording == "true" else "disabled"
            self.logger.info(f"✅ Azure AI Foundry tracing initialized successfully (content recording: {content_status}).")
            return True
            
        except Exception as e:
            self.logger.warning(f"⚠️ Failed to initialize tracing: {str(e)}")
            return False
    
    def setup_ui(self):
        """Setup the main UI layout."""
        # Create main paned window
        main_paned = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))
        
        # Left panel
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)
        
        # Right panel
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=2)
        
        self.setup_left_panel(left_frame)
        self.setup_right_panel(right_frame)
        # Note: status bar setup is called after StringVars are created
        
    def setup_left_panel(self, parent):
        """Setup the left panel with input controls."""
        # Context section
        context_label = ttk.Label(parent, text="Context")
        context_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.context_entry = tk.Entry(parent, width=40)
        self.context_entry.insert(0, "Microsoft Azure AI")  # Default value
        self.context_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Character Limit section
        limit_label = ttk.Label(parent, text="Character Limit")
        limit_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.limit_entry = tk.Entry(parent, width=40)
        self.limit_entry.insert(0, "2000")  # Default value
        self.limit_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Maximum Retries section
        retries_label = ttk.Label(parent, text="Maximum Retries")
        retries_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.retries_entry = tk.Entry(parent, width=40)
        self.retries_entry.insert(0, str(self.max_retries))  # Default value from constructor
        self.retries_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Question section
        question_label = ttk.Label(parent, text="Question")
        question_label.pack(anchor=tk.W, pady=(0, 5))
        
        self.question_text = scrolledtext.ScrolledText(parent, height=8, width=40, 
                                                     font=('Segoe UI', 12), wrap=tk.WORD)
        self.question_text.insert(tk.END, "Does your service offer video generative AI?")  # Default value
        self.question_text.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Buttons
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.ask_button = tk.Button(button_frame, text="Ask!", bg="lightgray", fg="black", 
                                  height=2, command=self.on_ask_clicked)
        self.ask_button.pack(fill=tk.X, pady=(0, 10))
        
        self.import_button = tk.Button(button_frame, text="📊 Import From Excel", 
                                     command=self.on_import_excel_clicked)
        self.import_button.pack(fill=tk.X)
        
    def setup_right_panel(self, parent):
        """Setup the right panel with output sections."""
        # Create notebook for tabbed interface
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Answer tab
        answer_frame = ttk.Frame(self.notebook)
        self.notebook.add(answer_frame, text="Answer")
        
        answer_label = ttk.Label(answer_frame, text="Answer")
        answer_label.pack(anchor=tk.W, pady=(5, 5))
        
        self.answer_text = scrolledtext.ScrolledText(answer_frame, wrap=tk.WORD, 
                                                   font=('Segoe UI', 12))
        self.answer_text.insert(tk.END, "Response will appear here after clicking Ask!")
        self.answer_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
        # Documentation tab
        docs_frame = ttk.Frame(self.notebook)
        self.notebook.add(docs_frame, text="Documentation")
        
        docs_label = ttk.Label(docs_frame, text="Documentation")
        docs_label.pack(anchor=tk.W, pady=(5, 5))
        
        self.docs_text = scrolledtext.ScrolledText(docs_frame, wrap=tk.WORD, 
                                                  font=('Segoe UI', 12))
        self.docs_text.insert(tk.END, "Documentation will appear here...")
        self.docs_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
        # Reasoning tab
        reasoning_frame = ttk.Frame(self.notebook)
        self.notebook.add(reasoning_frame, text="Reasoning")
        
        reasoning_label = ttk.Label(reasoning_frame, text="Reasoning")
        reasoning_label.pack(anchor=tk.W, pady=(5, 5))
        
        self.reasoning_text = scrolledtext.ScrolledText(reasoning_frame, wrap=tk.WORD, 
                                                      font=('Segoe UI', 12))
        self.reasoning_text.insert(tk.END, "Reasoning will appear here...")
        self.reasoning_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
    def setup_status_bar(self):
        """Setup the status bar at the bottom of the window."""
        # Create status bar frame
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, padx=10, pady=(5, 10))
        
        # Add separator line above status bar
        separator = ttk.Separator(status_frame, orient=tk.HORIZONTAL)
        separator.pack(fill=tk.X, pady=(0, 5))
        
        # Create main status bar content
        status_content = ttk.Frame(status_frame)
        status_content.pack(fill=tk.X)
        
        # Left side: Working status and current agent
        left_frame = ttk.Frame(status_content)
        left_frame.pack(side=tk.LEFT, padx=(0, 10))
        
        # Status indicator
        status_label = ttk.Label(left_frame, text="Status:")
        status_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_working_label = ttk.Label(left_frame, textvariable=self.status_working, 
                                              font=('Segoe UI', 9, 'bold'))
        self.status_working_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Current agent
        agent_label = ttk.Label(left_frame, text="Agent:")
        agent_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_agent_label = ttk.Label(left_frame, textvariable=self.status_agent,
                                            font=('Segoe UI', 9))
        self.status_agent_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Elapsed time
        time_label = ttk.Label(left_frame, text="Time:")
        time_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_time_label = ttk.Label(left_frame, textvariable=self.status_time,
                                           font=('Segoe UI', 9, 'bold'))
        self.status_time_label.pack(side=tk.LEFT)
        
        # Right side: Excel information (only shown when processing Excel)
        self.excel_frame = ttk.Frame(status_content)
        # Excel frame is packed/unpacked dynamically based on mode
        
        # Excel input file
        input_label = ttk.Label(self.excel_frame, text="Input:")
        input_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_excel_input_label = ttk.Label(self.excel_frame, textvariable=self.status_excel_input,
                                                  font=('Segoe UI', 9))
        self.status_excel_input_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Excel output file
        output_label = ttk.Label(self.excel_frame, text="Output:")
        output_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_excel_output_label = ttk.Label(self.excel_frame, textvariable=self.status_excel_output,
                                                   font=('Segoe UI', 9))
        self.status_excel_output_label.pack(side=tk.LEFT, padx=(0, 15))
        
        # Current question number
        question_label = ttk.Label(self.excel_frame, text="Question:")
        question_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.status_excel_question_label = ttk.Label(self.excel_frame, textvariable=self.status_excel_question,
                                                     font=('Segoe UI', 9, 'bold'))
        self.status_excel_question_label.pack(side=tk.LEFT)
        
    def start_working(self, agent_name=""):
        """Start working mode - begin timer and update status."""
        if self.headless_mode or not self.status_working:
            return
            
        self.status_working.set("Working")
        self.status_working_label.config(foreground="green")
        self.status_agent.set(agent_name)
        
        # Start timer
        import time
        self.start_time = time.time()
        self.update_timer()
        
    def stop_working(self):
        """Stop working mode - stop timer and update status."""
        if self.headless_mode or not self.status_working:
            return
            
        self.status_working.set("Idle")
        self.status_working_label.config(foreground="black")
        self.status_agent.set("")
        
        # Stop timer
        if self.timer_job:
            self.root.after_cancel(self.timer_job)
            self.timer_job = None
        self.start_time = None
        self.status_time.set("00:00")
        
    def update_agent(self, agent_name):
        """Update the current agent being processed."""
        if self.headless_mode or not self.status_agent:
            return
        self.status_agent.set(agent_name)
        
    def update_timer(self):
        """Update the elapsed time display."""
        if self.headless_mode or not self.start_time or not self.status_time:
            return
            
        import time
        elapsed = int(time.time() - self.start_time)
        minutes = elapsed // 60
        seconds = elapsed % 60
        self.status_time.set(f"{minutes:02d}:{seconds:02d}")
        
        # Schedule next update
        self.timer_job = self.root.after(1000, self.update_timer)
        
    def show_excel_mode(self, input_path, output_path):
        """Show Excel processing information in status bar."""
        if self.headless_mode or not self.status_excel_input:
            return
            
        # Show file paths (truncate if too long)
        import os
        input_name = os.path.basename(input_path)
        output_name = os.path.basename(output_path)
        
        self.status_excel_input.set(input_name)
        self.status_excel_output.set(output_name)
        self.status_excel_question.set("")
        
        # Pack the Excel frame to show it
        self.excel_frame.pack(side=tk.RIGHT, padx=(10, 0))
        
    def hide_excel_mode(self):
        """Hide Excel processing information from status bar."""
        if self.headless_mode or not self.status_excel_input:
            return
            
        # Clear Excel variables
        self.status_excel_input.set("")
        self.status_excel_output.set("")
        self.status_excel_question.set("")
        
        # Unpack the Excel frame to hide it
        self.excel_frame.pack_forget()
        
    def update_excel_question(self, question_number, total_questions=None):
        """Update the current question number being processed."""
        if self.headless_mode or not self.status_excel_question:
            return
            
        if total_questions:
            self.status_excel_question.set(f"{question_number}/{total_questions}")
        else:
            self.status_excel_question.set(str(question_number))
        
    def log_reasoning(self, message: str):
        """Add a message to the reasoning text area or CLI output."""
        if self.headless_mode:
            self.cli_output.append(message)
        else:
            self.reasoning_text.insert(tk.END, f"{message}\n")
            self.reasoning_text.see(tk.END)
            self.root.update_idletasks()
        
    def on_ask_clicked(self):
        """Handle the Ask button click."""
        # Disable button during processing
        self.ask_button.config(state=tk.DISABLED)
        
        # Switch to Reasoning tab immediately
        self.notebook.select(2)  # Index 2 is the Reasoning tab (Answer=0, Documentation=1, Reasoning=2)
        
        # Clear previous results
        self.answer_text.delete(1.0, tk.END)
        self.docs_text.delete(1.0, tk.END)
        self.reasoning_text.delete(1.0, tk.END)
        
        # Start working mode
        self.start_working("Initializing")
        self.hide_excel_mode()  # Ensure Excel mode is hidden for single questions
        
        # Get input values
        question = self.question_text.get(1.0, tk.END).strip()
        context = self.context_entry.get().strip()
        char_limit = int(self.limit_entry.get()) if self.limit_entry.get().isdigit() else 2000
        max_retries = int(self.retries_entry.get()) if self.retries_entry.get().isdigit() else self.max_retries
        
        if not question:
            messagebox.showwarning("Input Required", "Please enter a question.")
            self.ask_button.config(state=tk.NORMAL)
            self.stop_working()
            return
            
        # Run processing in separate thread
        thread = threading.Thread(target=self.process_single_question, 
                                args=(question, context, char_limit, max_retries))
        thread.daemon = True
        thread.start()
        
    def process_single_question(self, question: str, context: str, char_limit: int, max_retries: int):
        """Process a single question using the three-agent workflow."""
        try:
            # Use custom span for the entire workflow
            if self.tracer:
                with self.tracer.start_as_current_span("questionnaire_multi_agent_workflow") as span:
                    span.set_attribute("workflow.name", "Questionnaire Multi-Agent")
                    span.set_attribute("workflow.context", context)
                    span.set_attribute("workflow.char_limit", char_limit)
                    span.set_attribute("workflow.max_retries", max_retries)
                    span.set_attribute("question.text", question[:100] + "..." if len(question) > 100 else question)
                    success, answer, links = self._execute_workflow(question, context, char_limit, max_retries)
                    return success, answer, links
            else:
                success, answer, links = self._execute_workflow(question, context, char_limit, max_retries)
                return success, answer, links
                
        except Exception as e:
            self.logger.error(f"Error processing question: {e}")
            error_msg = str(e)  # Capture the error message as a string
            if not self.headless_mode:
                self.root.after(0, lambda: messagebox.showerror("Error", f"An error occurred: {error_msg}"))
            return False, f"Error: {error_msg}", []
        finally:
            # Re-enable button and stop working mode
            if not self.headless_mode:
                self.root.after(0, lambda: self.ask_button.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.stop_working())
    
    def _execute_workflow(self, question: str, context: str, char_limit: int, max_retries: int) -> Tuple[bool, str, List[str]]:
        """Internal method to execute the multi-agent workflow."""
        self.log_reasoning("=" * 80)
        self.log_reasoning(f"STARTING NEW QUESTION WORKFLOW")
        self.log_reasoning(f"Question: {question[:100]}{'...' if len(question) > 100 else ''}")
        self.log_reasoning(f"Context: {context}")
        self.log_reasoning(f"Character limit: {char_limit}")
        self.log_reasoning(f"Max retries: {max_retries}")
        self.log_reasoning("Context is FRESH (no previous attempt history)")
        self.log_reasoning("=" * 80)
        
        # Create agents if not already created
        if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
            self.create_agents()
        
        attempt = 1
        max_attempts = max_retries
        
        # Track valid links across retries for this question
        accumulated_valid_links = []
        
        # Track previous attempts and feedback for context improvement
        attempt_history = []
        
        # Track if we have a valid answer that just needs links (issue #10)
        validated_answer = None
        skip_answer_checker = False
        
        while attempt <= max_attempts:
            self.log_reasoning(f"Attempt {attempt}/{max_attempts}")
            
            # Step 1: Generate answer
            self.log_reasoning("Question Answerer: Generating answer...")
            candidate_answer, doc_urls = self.generate_answer(question, context, char_limit, attempt_history)
            
            if not candidate_answer:
                self.log_reasoning("Question Answerer failed to generate an answer")
                break
            
            # Log the raw answer before processing
            self.log_reasoning("=== RAW ANSWER FROM QUESTION ANSWERER ===")
            self.log_reasoning(candidate_answer)
            self.log_reasoning("=== END RAW ANSWER ===")
            
            # Remove links and citations, save links for documentation
            clean_answer, text_links = self.extract_links_and_clean(candidate_answer)
            
            self.log_reasoning("=== CLEANED ANSWER AFTER URL EXTRACTION ===")
            self.log_reasoning(clean_answer)
            self.log_reasoning("=== END CLEANED ANSWER ===")
            
            self.log_reasoning(f"Extracted {len(text_links)} URLs from answer text")
            for link in text_links:
                self.log_reasoning(f"  Text URL: {link}")
            
            # Combine documentation URLs from run steps with any URLs found in text
            all_links = list(set(doc_urls + text_links))  # Remove duplicates
            self.log_reasoning(f"Total combined URLs: {len(all_links)}")
            for link in all_links:
                self.log_reasoning(f"  Combined URL: {link}")
            
            # Check character limit
            if len(clean_answer) > char_limit:
                rejection_reason = f"Answer exceeds character limit ({len(clean_answer)} > {char_limit})"
                self.log_reasoning(f"{rejection_reason}. Retrying...")
                
                # Add to attempt history for next iteration
                attempt_history.append({
                    'attempt': attempt,
                    'answer': clean_answer,
                    'rejection_reason': rejection_reason,
                    'rejected_by': 'Character Limit Check'
                })
                
                attempt += 1
                continue
            
            # Step 2: Validate answer (skip if we already have a validated answer)
            if skip_answer_checker and validated_answer:
                self.log_reasoning("Answer Checker: Skipping validation - we already have a validated answer")
                answer_valid = True
                clean_answer = validated_answer  # Use the previously validated answer
            else:
                answer_valid, answer_feedback = self.validate_answer(question, clean_answer)
                
                if not answer_valid:
                    self.log_reasoning(f"Answer Checker rejected: {answer_feedback}")
                    
                    # Add to attempt history for next iteration
                    attempt_history.append({
                        'attempt': attempt,
                        'answer': clean_answer,
                        'rejection_reason': answer_feedback,
                        'rejected_by': 'Answer Checker'
                    })
                    
                    attempt += 1
                    continue
                else:
                    # Answer Checker approved - save this as our validated answer
                    validated_answer = clean_answer
            
            # Step 3: Validate links
            self.log_reasoning("Link Checker: Verifying URLs...")
            links_valid, valid_links, link_feedback = self.validate_links(all_links)
            
            # Accumulate any valid links found in this attempt
            if valid_links:
                for link in valid_links:
                    if link not in accumulated_valid_links:
                        accumulated_valid_links.append(link)
                        self.log_reasoning(f"Added valid link to accumulated collection: {link}")
            
            # Check if we have a valid answer and at least some valid links (current or accumulated)
            if not links_valid and not accumulated_valid_links:
                # No valid links in current attempt and no accumulated links from previous attempts
                self.log_reasoning(f"Link Checker rejected: {link_feedback}")
                
                # Issue #10 fix: If we have a validated answer but no links, 
                # don't throw away the answer - instead ask for links that support it
                if validated_answer and not skip_answer_checker:
                    self.log_reasoning("Answer is validated but has no links")
                    self.log_reasoning("Next attempt will keep the validated answer and ask for supporting links")
                    
                    # Add special context for next iteration to find supporting links
                    attempt_history.append({
                        'attempt': attempt,
                        'answer': validated_answer,
                        'rejection_reason': f"Good answer but needs supporting links: {link_feedback}",
                        'rejected_by': 'Link Checker (needs supporting links)',
                        'special_instruction': 'keep_answer_find_links'
                    })
                    
                    # Set flag to skip Answer Checker on next iteration
                    skip_answer_checker = True
                else:
                    # Standard rejection - no validated answer yet
                    attempt_history.append({
                        'attempt': attempt,
                        'answer': clean_answer,
                        'rejection_reason': link_feedback,
                        'rejected_by': 'Link Checker'
                    })
                
                attempt += 1
                continue
            elif not links_valid and accumulated_valid_links:
                # Current attempt has no valid links, but we have accumulated valid links from previous attempts
                self.log_reasoning(f"Link Checker found no valid links in current attempt, but reusing {len(accumulated_valid_links)} valid links from previous attempts")
                final_valid_links = accumulated_valid_links.copy()
            else:
                # Current attempt has valid links
                final_valid_links = accumulated_valid_links.copy()  # Use all accumulated links
            
            # All checks passed
            self.log_reasoning("All agents approved the answer!")
            self.log_reasoning(f"Final answer will use {len(final_valid_links)} documentation links")
            self.log_reasoning("=" * 80)
            self.log_reasoning("QUESTION WORKFLOW COMPLETED SUCCESSFULLY")
            self.log_reasoning("Context will be CLEARED for next question (no carryover of attempt history)")
            self.log_reasoning("=" * 80)
            
            # Update UI with results (only in non-headless mode)
            if not self.headless_mode:
                self.root.after(0, lambda: self.update_results(clean_answer, final_valid_links))
            
            # Return success with results
            return True, clean_answer, final_valid_links
            
        # Max attempts reached
        self.log_reasoning(f"Failed to generate acceptable answer after {max_attempts} attempts")
        if not self.headless_mode:
            self.root.after(0, lambda: messagebox.showerror("Processing Failed", 
                f"Could not generate an acceptable answer after {max_attempts} attempts."))
        
        return False, f"Failed to generate acceptable answer after {max_attempts} attempts", []
                
    def update_results(self, answer: str, links: List[str]):
        """Update the UI with the final answer and documentation."""
        self.answer_text.delete(1.0, tk.END)
        self.answer_text.insert(tk.END, answer)
        
        self.docs_text.delete(1.0, tk.END)
        if links:
            self.docs_text.insert(tk.END, "Documentation links:\n\n")
            for link in links:
                self.docs_text.insert(tk.END, f"• {link}\n")
        else:
            self.docs_text.insert(tk.END, "No documentation links found.")
    
    def update_question_display(self, question: str):
        """Update the Question box with the current question being processed."""
        if not self.headless_mode and self.question_text:
            self.question_text.delete(1.0, tk.END)
            self.question_text.insert(tk.END, question)
            self.root.update_idletasks()
    
    def cleanup_agents(self):
        """Clean up all Azure AI Foundry agent resources."""
        # Skip cleanup in mock mode or if already cleaned up
        if self.mock_mode:
            return
        
        try:
            # Clean up question answerer session
            if self.question_answerer_session:
                try:
                    self.question_answerer_session.__exit__(None, None, None)
                    self.logger.info("Cleaned up Question Answerer agent")
                except Exception as e:
                    self.logger.warning(f"Error cleaning up Question Answerer agent: {e}")
                finally:
                    self.question_answerer_session = None
                    self.question_answerer_id = None
            
            # Clean up answer checker session
            if self.answer_checker_session:
                try:
                    self.answer_checker_session.__exit__(None, None, None)
                    self.logger.info("Cleaned up Answer Checker agent")
                except Exception as e:
                    self.logger.warning(f"Error cleaning up Answer Checker agent: {e}")
                finally:
                    self.answer_checker_session = None
                    self.answer_checker_id = None
            
            # Clean up link checker session
            if self.link_checker_session:
                try:
                    self.link_checker_session.__exit__(None, None, None)
                    self.logger.info("Cleaned up Link Checker agent")
                except Exception as e:
                    self.logger.warning(f"Error cleaning up Link Checker agent: {e}")
                finally:
                    self.link_checker_session = None
                    self.link_checker_id = None
                    
            self.logger.info("Agent cleanup completed")
            
        except Exception as e:
            self.logger.error(f"Error during agent cleanup: {e}")
            
    def create_agents(self):
        """Create the three Azure AI Foundry agents using FoundryAgentSession for proper cleanup."""
        # Skip agent creation in mock mode
        if self.mock_mode:
            self.log_reasoning("Mock mode enabled - skipping Azure agent creation")
            self.question_answerer_id = "mock_question_answerer"
            self.answer_checker_id = "mock_answer_checker"
            self.link_checker_id = "mock_link_checker"
            return
            
        # Clean up any existing agents first
        self.cleanup_agents()
            
        try:
            self.log_reasoning("Creating agents...")
            
            # Get model deployment name from environment
            model_deployment = os.getenv("AZURE_OPENAI_MODEL_DEPLOYMENT")
            bing_resource_name = os.getenv("BING_CONNECTION_ID")
            
            if not bing_resource_name:
                raise ValueError("BING_CONNECTION_ID not found in environment variables")
            
            # Get the actual connection ID from the resource name
            try:
                connection = self.project_client.connections.get(name=bing_resource_name)
                conn_id = connection.id
            except Exception as conn_error:
                self.log_reasoning(f"ERROR: Failed to get Bing connection '{bing_resource_name}': {conn_error}")
                
                # List all available connections to help debug
                try:
                    self.log_reasoning("Listing all available connections:")
                    connections = self.project_client.connections.list()
                    for conn in connections:
                        self.log_reasoning(f"  - Name: '{conn.name}', Type: {getattr(conn, 'connection_type', 'unknown')}, ID: {conn.id}")
                    
                    if not connections:
                        self.log_reasoning("  No connections found in this project.")
                except Exception as list_error:
                    self.log_reasoning(f"  Could not list connections: {list_error}")
                
                raise ValueError(f"Bing connection '{bing_resource_name}' not found. Check BING_CONNECTION_ID in your .env file.") from conn_error
            
            # Create Bing grounding tool
            bing_tool = BingGroundingTool(connection_id=conn_id)
            
            # Create Question Answerer agent using FoundryAgentSession
            try:
                self.question_answerer_session = FoundryAgentSession(
                    client=self.project_client,
                    model=model_deployment,
                    name="Question Answerer",
                    instructions="You are a question answering agent. You MUST search the web extensively for evidence and synthesize accurate answers. Your answer must be based on current web search results. IMPORTANT: You must include the actual source URLs directly in your answer text. Write the full URLs (like https://docs.microsoft.com/example) in your response text where you reference information. Do not use citation markers like [1], (source), or 【†source】 - instead include the actual URLs, which you should always put at the end of your response, separated by newlines with no other text or formatting. Write in plain text without formatting. Your answer must end with a period and contain only complete sentences. Do not include any closing phrases like 'Learn more:', 'References:', 'For more information, see:', 'For more details, see:', 'Learn more at:', 'More information:', 'Additional resources:', or any similar calls-to-action at the end. There should only be prose, followed by a list of URLs for reference separated by newlines. Those URLs should be the ones provided by Bing. Always use the Bing grounding tool to search for current information.",
                    agent_config={"tools": bing_tool.definitions}
                )
                question_answerer, _ = self.question_answerer_session.__enter__()
                self.question_answerer_id = self.question_answerer_session.get_agent_id()
            except Exception as e:
                self.log_reasoning(f"ERROR: Failed to create Question Answerer agent with model '{model_deployment}': {e}")
                raise ValueError(f"Model deployment '{model_deployment}' not found. Check AZURE_OPENAI_MODEL_DEPLOYMENT in your .env file.") from e
            
            # Create Answer Checker agent using FoundryAgentSession
            self.answer_checker_session = FoundryAgentSession(
                client=self.project_client,
                model=model_deployment,
                name="Answer Checker",
                instructions="You are an answer validation agent. Review candidate answers for factual correctness, completeness, and consistency. Use web search to verify claims. Respond with 'VALID' if the answer is acceptable or 'INVALID: [reason]' if not.",
                agent_config={"tools": bing_tool.definitions}
            )
            answer_checker, _ = self.answer_checker_session.__enter__()
            self.answer_checker_id = self.answer_checker_session.get_agent_id()
            
            # Create Link Checker agent using FoundryAgentSession
            self.link_checker_session = FoundryAgentSession(
                client=self.project_client,
                model=model_deployment,
                name="Link Checker",
                instructions="You are a link validation agent. Verify that URLs are reachable and relevant to the given question. Report any issues with links.",
                agent_config={"tools": []}  # Will use requests/playwright for link checking
            )
            link_checker, _ = self.link_checker_session.__enter__()
            self.link_checker_id = self.link_checker_session.get_agent_id()
            
            self.log_reasoning("All agents created successfully!")
            
        except Exception as e:
            self.logger.error(f"Failed to create agents: {e}")
            # Clean up any partially created agents
            self.cleanup_agents()
            raise
            
    def generate_answer(self, question: str, context: str, char_limit: int, attempt_history: list = None) -> Tuple[Optional[str], List[str]]:
        """Generate an answer using the Question Answerer agent."""
        try:
            # Use custom span for Question Answerer agent
            if self.tracer:
                with self.tracer.start_as_current_span("question_answerer_agent") as span:
                    span.set_attribute("agent.name", "Question Answerer")
                    span.set_attribute("agent.operation", "generate_answer")
                    span.set_attribute("question.context", context)
                    span.set_attribute("question.char_limit", char_limit)
                    return self._execute_question_answerer(question, context, char_limit, attempt_history)
            else:
                return self._execute_question_answerer(question, context, char_limit, attempt_history)
                
        except Exception as e:
            error_msg = f"Error generating answer: {e}"
            self.logger.error(error_msg)
            self.log_reasoning(error_msg)
            return None, []
    
    def _execute_question_answerer(self, question: str, context: str, char_limit: int, attempt_history: list = None) -> Tuple[Optional[str], List[str]]:
        """Internal method to execute Question Answerer agent operations."""
        # Check for mock mode first
        if self.mock_mode:
            return self._execute_question_answerer_mock(question, context, char_limit, attempt_history)
        
        # Update status bar
        self.update_agent("Question Answerer")
        
        # Create thread
        thread = self.project_client.agents.threads.create()
        
        # Create message
        prompt_content = f"Context: {context}\n\nQuestion: {question}\n\n"
        
        # Add previous attempt history if available
        if attempt_history:
            self.log_reasoning(f"Question Answerer: Modifying context with {len(attempt_history)} previous attempt(s)")
            
            # Check if the last attempt has special instructions (Issue #10)
            last_attempt = attempt_history[-1] if attempt_history else None
            if last_attempt and last_attempt.get('special_instruction') == 'keep_answer_find_links':
                self.log_reasoning("Question Answerer: Special instruction - keep validated answer and find supporting links")
                prompt_content += "SPECIAL INSTRUCTION:\n"
                prompt_content += f"The following answer has been validated as factually correct and complete:\n\n"
                prompt_content += f'"{last_attempt["answer"]}"\n\n'
                prompt_content += "DO NOT rewrite this answer. Instead, use your web search capabilities to find credible sources and documentation links that support the claims made in this answer. "
                prompt_content += "Return the EXACT same answer text, but include the actual source URLs that you find to substantiate the information. "
                prompt_content += "Put the URLs at the end of your response, separated by newlines.\n\n"
                self.log_reasoning("Question Answerer: Added special instruction to keep answer and find supporting links")
            else:
                # Standard attempt history processing
                prompt_content += "PREVIOUS ATTEMPTS AND FEEDBACK:\n"
                for i, attempt_info in enumerate(attempt_history, 1):
                    prompt_content += f"\nAttempt {attempt_info['attempt']}:\n"
                    prompt_content += f"Previous Answer: {attempt_info['answer'][:500]}{'...' if len(attempt_info['answer']) > 500 else ''}\n"
                    prompt_content += f"Rejected by {attempt_info['rejected_by']}: {attempt_info['rejection_reason']}\n"
                    self.log_reasoning(f"Question Answerer: Added feedback from attempt {attempt_info['attempt']} (rejected by {attempt_info['rejected_by']})")
                prompt_content += "\nBased on the above feedback, please generate an improved answer that addresses the identified issues.\n\n"
                self.log_reasoning("Question Answerer: Context enhanced with previous attempt feedback")
        else:
            self.log_reasoning("Question Answerer: Using original context (no previous attempts)")
        
        prompt_content += f"Please provide a comprehensive answer with supporting evidence and citations. Keep it under {char_limit} characters."
        message = self.project_client.agents.messages.create(
            thread_id=thread.id,
            role="user",
            content=prompt_content
        )
        
        # Add span attributes for better metrics capture
        if self.tracer:
            current_span = trace.get_current_span()
            if current_span:
                # Add generative AI semantic convention attributes
                current_span.set_attribute("gen_ai.system", "azure_ai_foundry")
                current_span.set_attribute("gen_ai.operation.name", "chat_completion")
                current_span.set_attribute("gen_ai.request.model", os.getenv("AZURE_OPENAI_MODEL_DEPLOYMENT", "unknown"))
                current_span.set_attribute("gen_ai.usage.prompt_tokens", len(prompt_content.split()))  # Rough estimate
                current_span.set_attribute("thread.id", thread.id)
                current_span.set_attribute("agent.id", self.question_answerer_id)
        
        # Create and process run
        run = self.project_client.agents.runs.create_and_process(
            thread_id=thread.id,
            agent_id=self.question_answerer_id
        )
        
        self.log_reasoning(f"Run finished with status: {run.status}")
        
        # Update span with completion attributes
        if self.tracer:
            current_span = trace.get_current_span()
            if current_span:
                current_span.set_attribute("run.status", run.status)
                current_span.set_attribute("run.id", run.id)
        
        # Check if the run failed
        if run.status == "failed":
            error_msg = f"Run failed: {run.last_error.message if run.last_error else 'Unknown error'}"
            self.log_reasoning(error_msg)
            self.logger.error(error_msg)
            if self.tracer:
                current_span = trace.get_current_span()
                if current_span:
                    current_span.set_attribute("error", True)
                    current_span.set_attribute("error.message", error_msg)
            return None, []
        
        # Get messages
        messages = self.project_client.agents.messages.list(thread_id=thread.id)
        
        # Find the assistant's response
        for msg in messages:
            if msg.role == "assistant" and msg.content:
                response = msg.content[0].text.value
                self.log_reasoning(f"Got response: {response[:100]}...")
                
                # Extract actual source URLs from message annotations
                doc_urls = []
                if hasattr(msg.content[0], 'annotations') and msg.content[0].annotations:
                    self.log_reasoning(f"Found {len(msg.content[0].annotations)} annotations")
                    for annotation in msg.content[0].annotations:
                        if hasattr(annotation, 'uri_citation') and annotation.uri_citation:
                            url = annotation.uri_citation.uri
                            # Only include actual website URLs, not Bing API URLs
                            if url and not url.startswith('https://api.bing.microsoft.com'):
                                doc_urls.append(url)
                                self.log_reasoning(f"Found source URL: {url}")
                
                if not doc_urls:
                    self.log_reasoning("No annotations found, trying run steps...")
                    doc_urls = self.extract_documentation_urls(thread.id, run.id)
                
                return response, doc_urls
                
        self.log_reasoning("No assistant response found in messages")
        return None, []
            
    def extract_documentation_urls(self, thread_id: str, run_id: str) -> List[str]:
        """Extract documentation URLs from run steps."""
        try:
            run_steps = self.project_client.agents.run_steps.list(thread_id=thread_id, run_id=run_id)
            documentation_urls = []
            
            for step in run_steps:
                
                # Check if there are tool calls in the step details
                if hasattr(step, 'step_details') and step.step_details:
                    step_details = step.step_details
                    if hasattr(step_details, 'tool_calls') and step_details.tool_calls:
                        self.log_reasoning(f"Found {len(step_details.tool_calls)} tool calls")
                        
                        for call in step_details.tool_calls:
                            # Look for any tool call output that might contain URLs
                            if hasattr(call, 'bing_grounding') and call.bing_grounding:
                                self.log_reasoning(f"Examining bing_grounding data: {type(call.bing_grounding)}")
                                
                                # Handle different data structures
                                if isinstance(call.bing_grounding, dict):
                                    for key, value in call.bing_grounding.items():
                                        self.log_reasoning(f"  Key: {key}, Value type: {type(value)}")
                                        if isinstance(value, str) and value.startswith('http') and not value.startswith('https://api.bing.microsoft.com'):
                                            documentation_urls.append(value)
                                            self.log_reasoning(f"Found documentation URL from dict: {value}")
                                elif isinstance(call.bing_grounding, list):
                                    for item in call.bing_grounding:
                                        if isinstance(item, dict):
                                            for key, value in item.items():
                                                if isinstance(value, str) and value.startswith('http') and not value.startswith('https://api.bing.microsoft.com'):
                                                    documentation_urls.append(value)
                                                    self.log_reasoning(f"Found documentation URL from list: {value}")
                                        elif isinstance(item, str) and item.startswith('http') and not item.startswith('https://api.bing.microsoft.com'):
                                            documentation_urls.append(item)
                                            self.log_reasoning(f"Found documentation URL from list item: {item}")
            
            if documentation_urls:
                self.log_reasoning(f"Extracted {len(documentation_urls)} documentation URLs")
            else:
                self.log_reasoning("No documentation URLs found in run steps")
                
            return list(set(documentation_urls))  # Remove duplicates
            
        except Exception as e:
            self.log_reasoning(f"Error extracting documentation URLs: {e}")
            return []
            
    def validate_answer(self, question: str, answer: str) -> Tuple[bool, str]:
        """Validate an answer using the Answer Checker agent."""
        try:
            # Use custom span for Answer Checker agent
            if self.tracer:
                with self.tracer.start_as_current_span("answer_checker_agent") as span:
                    span.set_attribute("agent.name", "Answer Checker")
                    span.set_attribute("agent.operation", "validate_answer")
                    span.set_attribute("answer.length", len(answer))
                    return self._execute_answer_checker(question, answer)
            else:
                return self._execute_answer_checker(question, answer)
                
        except Exception as e:
            self.logger.error(f"Error validating answer: {e}")
            return False, f"Error: {e}"
    
    def _execute_answer_checker(self, question: str, answer: str) -> Tuple[bool, str]:
        """Internal method to execute Answer Checker agent operations."""
        # Check for mock mode first
        if self.mock_mode:
            return self._execute_answer_checker_mock(question, answer)
        
        # Update status bar
        self.update_agent("Answer Checker")
        
        # Create thread
        thread = self.project_client.agents.threads.create()
        
        # Create message
        prompt_content = f"Question: {question}\n\nCandidate Answer: {answer}\n\nPlease validate this answer for factual correctness, completeness, and consistency. Respond with 'VALID' if acceptable or 'INVALID: [reason]' if not."
        message = self.project_client.agents.messages.create(
            thread_id=thread.id,
            role="user",
            content=prompt_content
        )
        
        # Add span attributes for better metrics capture
        if self.tracer:
            current_span = trace.get_current_span()
            if current_span:
                current_span.set_attribute("gen_ai.system", "azure_ai_foundry")
                current_span.set_attribute("gen_ai.operation.name", "chat_completion")
                current_span.set_attribute("gen_ai.request.model", os.getenv("AZURE_OPENAI_MODEL_DEPLOYMENT", "unknown"))
                current_span.set_attribute("gen_ai.usage.prompt_tokens", len(prompt_content.split()))
                current_span.set_attribute("thread.id", thread.id)
                current_span.set_attribute("agent.id", self.answer_checker_id)
        
        # Create and process run
        run = self.project_client.agents.runs.create_and_process(
            thread_id=thread.id,
            agent_id=self.answer_checker_id
        )
        
        # Update span with completion attributes
        if self.tracer:
            current_span = trace.get_current_span()
            if current_span:
                current_span.set_attribute("run.status", run.status)
                current_span.set_attribute("run.id", run.id)
        
        # Get messages
        messages = self.project_client.agents.messages.list(thread_id=thread.id)
        
        # Find the assistant's response
        for msg in messages:
            if msg.role == "assistant" and msg.content:
                response = msg.content[0].text.value
                
                if "VALID" in response.upper() and "INVALID" not in response.upper():
                    self.log_reasoning("Answer Checker: APPROVED the answer")
                    return True, response
                else:
                    self.log_reasoning("Answer Checker: REJECTED the answer")
                    self.log_reasoning(f"Answer Checker: Rejection reason: {response}")
                    return False, response
                    
        self.log_reasoning("Answer Checker: ERROR - No response received")
        return False, "No response from Answer Checker"
            
    def validate_links(self, links: List[str]) -> Tuple[bool, List[str], str]:
        """Validate links using CURL and Link Checker agent."""
        try:
            # Use custom span for Link Checker agent
            if self.tracer:
                with self.tracer.start_as_current_span("link_checker_agent") as span:
                    span.set_attribute("agent.name", "Link Checker")
                    span.set_attribute("agent.operation", "validate_links")
                    span.set_attribute("links.count", len(links))
                    return self._execute_link_checker(links)
            else:
                return self._execute_link_checker(links)
                
        except Exception as e:
            self.logger.error(f"Error validating links: {e}")
            return False, [], f"Error: {e}"
    
    def _execute_link_checker(self, links: List[str]) -> Tuple[bool, List[str], str]:
        """Internal method to execute Link Checker agent operations."""
        # Check for mock mode first
        if self.mock_mode:
            return self._execute_link_checker_mock(links)
        
        # Update status bar
        self.update_agent("Link Checker")
        
        import requests
        
        # First check: Must have at least one URL
        if not links or len(links) == 0:
            self.log_reasoning("Link Checker: No documentation URLs found - this is a failure condition")
            return False, [], "No documentation URLs provided. All answers must include source links."
        
        self.log_reasoning(f"Link Checker: Validating {len(links)} URLs")
        
        valid_links = []
        invalid_links = []
        
        # Check if links are reachable using requests
        for link in links:
            try:
                response = requests.head(link, timeout=10, allow_redirects=True)
                if response.status_code == 200:
                    valid_links.append(link)
                    self.log_reasoning(f"Link Checker: ✓ {link} (HTTP 200)")
                else:
                    invalid_links.append(f"{link} (HTTP {response.status_code})")
                    self.log_reasoning(f"Link Checker: ✗ {link} (HTTP {response.status_code})")
            except Exception as e:
                invalid_links.append(f"{link} (Error: {e})")
                self.log_reasoning(f"Link Checker: ✗ {link} (Error: {e})")
        
        # If we have at least one valid link, that's success
        if valid_links:
            if invalid_links:
                self.log_reasoning(f"Link Checker: Removed {len(invalid_links)} invalid URLs, keeping {len(valid_links)} valid ones")
                return True, valid_links, f"Found {len(valid_links)} valid links (removed {len(invalid_links)} invalid)"
            else:
                return True, valid_links, f"All {len(valid_links)} links are valid"
        else:
            return False, [], "No valid documentation URLs found after validation"
    
    # Mock implementation methods for testing
    def _execute_question_answerer_mock(self, question: str, context: str, char_limit: int, attempt_history: list = None) -> Tuple[Optional[str], List[str]]:
        """Mock implementation of Question Answerer for testing."""
        self.log_reasoning("Question Answerer (MOCK): Generating mock response")
        
        # Generate a contextual mock answer
        mock_answer = f"Based on the {context} context, regarding '{question[:50]}...': "
        
        if "azure" in question.lower():
            mock_answer += "Microsoft Azure provides comprehensive cloud services including AI, compute, storage, and networking capabilities. "
        elif "ai" in question.lower() or "artificial intelligence" in question.lower():
            mock_answer += "Artificial Intelligence capabilities are available through various cloud platforms and services. "
        elif "video" in question.lower():
            mock_answer += "Video processing and generation capabilities are available through modern AI services. "
        else:
            mock_answer += "This is a comprehensive answer to your question with relevant information and supporting documentation. "
        
        mock_answer += "For detailed information and documentation, please refer to the official Microsoft resources."
        
        # Ensure we stay under the character limit
        if len(mock_answer) > char_limit:
            mock_answer = mock_answer[:char_limit-3] + "..."
        
        # Always include Microsoft.com as a documentation link
        mock_links = ["https://www.microsoft.com"]
        
        self.log_reasoning(f"Question Answerer (MOCK): Generated {len(mock_answer)} character response with {len(mock_links)} link(s)")
        
        return mock_answer, mock_links
    
    def _execute_answer_checker_mock(self, question: str, answer: str) -> Tuple[bool, str]:
        """Mock implementation of Answer Checker for testing."""
        self.log_reasoning("Answer Checker (MOCK): APPROVED the answer")
        return True, "VALID"
    
    def _execute_link_checker_mock(self, links: List[str]) -> Tuple[bool, List[str], str]:
        """Mock implementation of Link Checker for testing."""
        self.log_reasoning(f"Link Checker (MOCK): Validating {len(links)} URLs")
        
        # Microsoft.com should always be valid, so this will pass
        valid_links = []
        for link in links:
            if "microsoft.com" in link.lower():
                valid_links.append(link)
                self.log_reasoning(f"Link Checker (MOCK): ✓ {link} (HTTP 200)")
            else:
                # For testing purposes, let's assume other URLs are also valid in mock mode
                valid_links.append(link)
                self.log_reasoning(f"Link Checker (MOCK): ✓ {link} (Mock validation)")
        
        if valid_links:
            return True, valid_links, f"All {len(valid_links)} links are valid (mock validation)"
        else:
            # This should not happen with our mock Question Answerer, but handle it gracefully
            return False, [], "No links provided for validation"
    
    def identify_columns_mock(self, df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """Mock implementation of column identification for testing."""
        self.log_reasoning("Column Identification (MOCK): Analyzing columns...")
        
        # Simple heuristic-based column identification for mock mode
        columns = df.columns.tolist()
        question_col = None
        answer_col = None
        docs_col = None
        
        # Look for question-like columns
        for col in columns:
            col_lower = col.lower()
            if any(word in col_lower for word in ['question', 'q', 'query', 'ask']):
                question_col = col
                break
        
        # Look for answer-like columns with priority order
        # Priority: exact matches for Response/Answer/Responses/Answers, then other terms
        priority_terms = ['response', 'answer', 'responses', 'answers']
        secondary_terms = ['reply', 'result']
        
        # First, look for priority terms (exact matches)
        for col in columns:
            col_lower = col.lower()
            if col_lower in priority_terms:
                answer_col = col
                break
        
        # If no priority match found, look for columns containing priority terms
        if not answer_col:
            for col in columns:
                col_lower = col.lower()
                if any(term in col_lower for term in priority_terms):
                    answer_col = col
                    break
        
        # Finally, fall back to secondary terms
        if not answer_col:
            for col in columns:
                col_lower = col.lower()
                if any(term in col_lower for term in secondary_terms + ['a']):  # 'a' as last resort
                    answer_col = col
                    break
        
        # Look for documentation columns
        for col in columns:
            col_lower = col.lower()
            if any(word in col_lower for word in ['doc', 'link', 'ref', 'source', 'documentation', 'url']):
                docs_col = col
                break
        
        # If we couldn't find obvious matches, make educated guesses based on position and content
        if not question_col and len(columns) >= 1:
            # Check first few columns for question-like content
            for i, col in enumerate(columns[:3]):
                sample_data = df[col].dropna().head(3).astype(str).tolist()
                if any('?' in str(item) for item in sample_data):
                    question_col = col
                    break
        
        if not answer_col and len(columns) >= 2:
            # Look for an empty or mostly empty column that could be for answers
            for col in columns:
                if col != question_col:
                    empty_ratio = df[col].isna().sum() / len(df)
                    if empty_ratio > 0.5:  # More than 50% empty
                        answer_col = col
                        break
        
        # If still no answer column, use the second column if available
        if not answer_col and len(columns) >= 2:
            for col in columns:
                if col != question_col:
                    answer_col = col
                    break
        
        self.log_reasoning(f"Column Identification (MOCK): Question='{question_col}', Answer='{answer_col}', Docs='{docs_col}'")
        
        return question_col, answer_col, docs_col
        
    def extract_links_and_clean(self, text: str) -> Tuple[str, List[str]]:
        """Extract URLs from text and return cleaned text and list of URLs."""
        import re
        
        self.log_reasoning(f"Extracting URLs from text (length: {len(text)})")
        self.log_reasoning(f"Text preview: {text[:200]}...")
        
        # Find URLs
        url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
        urls = re.findall(url_pattern, text)
        self.log_reasoning(f"URL regex found {len(urls)} URLs")
        
        # Remove URLs from text
        clean_text = re.sub(url_pattern, '', text)
        
        # Remove various citation formats
        clean_text = re.sub(r'\[\d+\]', '', clean_text)  # Remove [1], [2], etc.
        clean_text = re.sub(r'【[^】]*】', '', clean_text)  # Remove 【3:3†source】 style citations
        clean_text = re.sub(r'\(\d+\)', '', clean_text)  # Remove (1), (2), etc.
        clean_text = re.sub(r'\[\d+:\d+[^]]*\]', '', clean_text)  # Remove [3:3†source] style
        
        # Remove markdown formatting
        clean_text = re.sub(r'\*\*(.*?)\*\*', r'\1', clean_text)  # Remove **bold**
        clean_text = re.sub(r'\*(.*?)\*', r'\1', clean_text)  # Remove *italic*
        clean_text = re.sub(r'`(.*?)`', r'\1', clean_text)  # Remove `code`
        clean_text = re.sub(r'#{1,6}\s*', '', clean_text)  # Remove # headers
        
        # Remove numbered list formatting
        clean_text = re.sub(r'^\d+\.\s*\*\*[^*]*\*\*:\s*', '', clean_text, flags=re.MULTILINE)  # Remove "1. **Title**:"
        clean_text = re.sub(r'^\d+\.\s*', '', clean_text, flags=re.MULTILINE)  # Remove "1. "
        
        # Remove bullet points
        clean_text = re.sub(r'^\s*[-•]\s*', '', clean_text, flags=re.MULTILINE)
        
        # Remove "References:" and similar closing phrases at the end
        clean_text = re.sub(r'\s*References?:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*For more information,?\s*see:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*For more information,?\s*visit:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*For more details,?\s*see:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*Learn more:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*Learn more at:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*More information:\s*$', '', clean_text, flags=re.IGNORECASE)
        clean_text = re.sub(r'\s*Additional resources:\s*$', '', clean_text, flags=re.IGNORECASE)
        
        # Clean up whitespace and line breaks
        clean_text = re.sub(r'\n\s*\n', ' ', clean_text)  # Replace multiple newlines with space
        clean_text = re.sub(r'\s+', ' ', clean_text).strip()  # Clean up multiple spaces
        
        # Fix space before periods (e.g., "these scenarios ." -> "these scenarios.")
        clean_text = re.sub(r'\s+\.', '.', clean_text)  # Replace space(s) before period with just period
        
        return clean_text, urls
        
    def on_import_excel_clicked(self):
        """Handle the Import From Excel button click."""
        # First, get the input Excel file
        input_file_path = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if not input_file_path:
            return
        
        # Then, get the output Excel file location
        output_file_path = filedialog.asksaveasfilename(
            title="Save Processed Excel File As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file_path:
            return
            
        # Switch to Reasoning tab immediately to show processing status
        self.notebook.select(2)  # Index 2 is the Reasoning tab (Answer=0, Documentation=1, Reasoning=2)
        
        # Clear reasoning text to show fresh processing logs
        self.reasoning_text.delete(1.0, tk.END)
        
        # Clear the question text box since we're starting Excel processing
        if self.question_text:
            self.question_text.delete(1.0, tk.END)
        
        # Start Excel processing mode
        self.start_working("Excel Processing")
        self.show_excel_mode(input_file_path, output_file_path)
        
        # Process Excel file in separate thread with both input and output paths
        thread = threading.Thread(target=self.process_excel_file, args=(input_file_path, output_file_path))
        thread.daemon = True
        thread.start()
            
    def process_excel_file(self, input_file_path: str, output_file_path: str):
        """Process an Excel file with questions, saving results continuously to output file."""
        try:
            self.log_reasoning(f"Processing Excel file: {input_file_path}")
            self.log_reasoning(f"Output will be saved to: {output_file_path}")
            
            # Create agents if not already created
            if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
                self.create_agents()
            
            # Create temporary file to work with
            import tempfile
            import shutil
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # Copy input file to temporary location
            shutil.copy2(input_file_path, temp_path)
            self.log_reasoning(f"Created temporary working file: {temp_path}")
            
            # Read Excel file
            # Suppress openpyxl data validation warnings when reading Excel metadata
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                excel_file = pd.ExcelFile(temp_path)
            context = self.context_entry.get().strip()
            char_limit = int(self.limit_entry.get()) if self.limit_entry.get().isdigit() else 2000
            max_retries = int(self.retries_entry.get()) if self.retries_entry.get().isdigit() else self.max_retries
            
            # Load workbook once for all sheets
            from openpyxl import load_workbook
            # Suppress openpyxl data validation warnings when loading Excel
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                wb = load_workbook(temp_path)
            
            for sheet_name in excel_file.sheet_names:
                self.log_reasoning(f"Processing sheet: {sheet_name}")
                # Suppress openpyxl data validation warnings when reading Excel
                with warnings.catch_warnings():
                    warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                    df = pd.read_excel(temp_path, sheet_name=sheet_name)
                
                # Use LLM to identify columns (reuse CLI logic)
                question_col, answer_col, docs_col = self.identify_columns_with_llm_cli(df)
                
                self.log_reasoning(f"LLM identified columns - Questions: {question_col}, Answers: {answer_col}, Docs: {docs_col}")
                
                # Skip sheet if no question or answer column found
                if not question_col or not answer_col:
                    self.log_reasoning(f"Skipping sheet '{sheet_name}' - missing required question or answer column")
                    continue
                
                # Ensure answer and documentation columns are properly typed for string assignment
                # This prevents pandas FutureWarning about incompatible dtype when setting string values
                if answer_col and answer_col in df.columns:
                    df[answer_col] = df[answer_col].astype('object')
                if docs_col and docs_col in df.columns:
                    df[docs_col] = df[docs_col].astype('object')
                
                # Get worksheet for this sheet
                ws = wb[sheet_name]
                
                # Count total questions in this sheet first
                total_questions_in_sheet = 0
                for _, row in df.iterrows():
                    if pd.notna(row[question_col]) and str(row[question_col]).strip():
                        total_questions_in_sheet += 1
                
                # Process each question
                questions_processed = 0
                questions_attempted = 0
                for idx, row in df.iterrows():
                    if pd.notna(row[question_col]) and str(row[question_col]).strip():
                        question = str(row[question_col]).strip()
                        questions_attempted += 1
                        self.log_reasoning(f"Processing question {idx + 1}: {question[:50]}...")
                        
                        # Update status bar with current question number
                        if not self.headless_mode:
                            self.root.after(0, lambda qnum=questions_attempted, qtotal=total_questions_in_sheet: 
                                           self.update_excel_question(qnum, qtotal))
                        
                        # Update the Question box to show current question being processed
                        if not self.headless_mode:
                            self.root.after(0, lambda q=question: self.update_question_display(q))
                        
                        # Process question using the same workflow as single questions
                        success, answer, links = self.process_question_with_agents(question, context, char_limit, max_retries)
                        
                        if success:
                            # Update the answer column in dataframe
                            df.at[idx, answer_col] = answer
                            
                            # Update documentation column only if it exists and we have links
                            if docs_col and links:
                                df.at[idx, docs_col] = '\n'.join(links)
                            
                            # Save to Excel file immediately using openpyxl
                            row_num = idx + 2  # +2 because Excel is 1-indexed and has header
                            
                            # Find and update answer column
                            for col_idx, col_name in enumerate(df.columns, 1):
                                if col_name == answer_col:
                                    cell = ws.cell(row=row_num, column=col_idx)
                                    cell.value = answer
                                elif col_name == docs_col and docs_col and links:
                                    cell = ws.cell(row=row_num, column=col_idx)
                                    cell.value = '\n'.join(links)
                            
                            # Save the workbook immediately after each successful question
                            wb.save(temp_path)
                            
                            self.log_reasoning(f"Successfully processed question {idx + 1} and saved to temporary file")
                            questions_processed += 1
                        else:
                            self.log_reasoning(f"Failed to process question {idx + 1}: {answer}")
                            # Leave response blank on failure - don't write error messages  
                            # Leave documentation blank on failure - don't write error messages
                
                self.log_reasoning(f"Processed {questions_processed}/{questions_attempted} questions successfully in sheet '{sheet_name}'")
            
            # Close the workbook after processing all sheets
            wb.close()
            
            # Close the Excel file
            excel_file.close()
            
            # Copy the completed temporary file to the final output location
            shutil.copy2(temp_path, output_file_path)
            self.log_reasoning(f"Copied completed file from {temp_path} to {output_file_path}")
            
            # Clean up temporary file
            try:
                os.unlink(temp_path)
                self.log_reasoning(f"Cleaned up temporary file: {temp_path}")
            except Exception as cleanup_error:
                self.logger.warning(f"Could not clean up temporary file {temp_path}: {cleanup_error}")
            
            self.log_reasoning(f"Excel processing completed. All results saved to: {output_file_path}")
            
            # Show completion message
            if not self.headless_mode:
                self.root.after(0, lambda: messagebox.showinfo("Success", f"File processed and saved to:\n{output_file_path}"))
                
        except Exception as e:
            self.logger.error(f"Error processing Excel file: {e}")
            
            # Clean up temporary file if it exists
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                    self.log_reasoning(f"Cleaned up temporary file after error: {temp_path}")
                except Exception as cleanup_error:
                    self.logger.warning(f"Could not clean up temporary file {temp_path} after error: {cleanup_error}")
            
            if not self.headless_mode:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to process Excel file:\n{e}"))
        finally:
            # Stop working mode and hide Excel mode
            if not self.headless_mode:
                self.root.after(0, lambda: self.stop_working())
                self.root.after(0, lambda: self.hide_excel_mode())
            
    def save_processed_excel(self, temp_path: str):
        """Save the processed Excel file."""
        try:
            save_path = filedialog.asksaveasfilename(
                title="Save Processed Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if save_path:
                import shutil
                shutil.move(temp_path, save_path)
                messagebox.showinfo("Success", f"File processed and saved to:\n{save_path}")
                
                # Verify that the file was successfully saved and the temporary file no longer exists
                if os.path.exists(save_path) and not os.path.exists(temp_path):
                    self.logger.info(f"Excel file successfully saved to: {save_path}")
                elif os.path.exists(save_path) and os.path.exists(temp_path):
                    # File was saved but temp file still exists - try to clean it up silently
                    try:
                        os.unlink(temp_path)
                        self.logger.info(f"Cleaned up temporary file: {temp_path}")
                    except Exception as cleanup_error:
                        # Log the cleanup error but don't show it to the user since the main operation succeeded
                        self.logger.warning(f"Could not clean up temporary file {temp_path}: {cleanup_error}")
                else:
                    # This shouldn't happen, but log it for debugging
                    self.logger.warning(f"Unexpected state after file save - save_path exists: {os.path.exists(save_path)}, temp_path exists: {os.path.exists(temp_path)}")
            else:
                # User cancelled the save dialog - clean up temp file silently
                try:
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)
                        self.logger.info(f"User cancelled save dialog, cleaned up temporary file: {temp_path}")
                    else:
                        self.logger.info("User cancelled save dialog, temporary file already cleaned up")
                except Exception as cleanup_error:
                    # Log cleanup errors but don't show them to the user since they cancelled the operation
                    self.logger.warning(f"Could not clean up temporary file after user cancellation {temp_path}: {cleanup_error}")
        except Exception as e:
            # Only show errors that actually prevent the file from being saved
            # Check if the save was actually successful despite the error
            if 'save_path' in locals() and save_path and os.path.exists(save_path):
                # File was saved successfully, so don't show the scary error dialog
                self.logger.warning(f"File was saved successfully to {save_path} but encountered error during cleanup: {e}")
                messagebox.showinfo("Success", f"File processed and saved to:\n{save_path}\n\n(Note: Encountered minor cleanup issue, but your file was saved successfully)")
            else:
                # Actual save failure - show the error
                self.logger.error(f"Error saving Excel file: {e}")
                messagebox.showerror("Error", f"Failed to save Excel file:\n{e}")
                
                # Try to clean up temp file even if save failed
                try:
                    if os.path.exists(temp_path):
                        os.unlink(temp_path)
                except Exception as cleanup_error:
                    self.logger.warning(f"Could not clean up temporary file after save failure {temp_path}: {cleanup_error}")
            


    def process_question_with_agents(self, question: str, context: str, char_limit: int, max_retries: int) -> Tuple[bool, str, List[str]]:
        """Process a single question using the three-agent workflow."""
        self.log_reasoning("=" * 80)
        self.log_reasoning(f"STARTING NEW QUESTION WORKFLOW (Excel Processing)")
        self.log_reasoning(f"Question: {question[:100]}{'...' if len(question) > 100 else ''}")
        self.log_reasoning(f"Context: {context}")
        self.log_reasoning(f"Character limit: {char_limit}")
        self.log_reasoning(f"Max retries: {max_retries}")
        self.log_reasoning("Context is FRESH (no previous attempt history)")
        self.log_reasoning("=" * 80)
        
        try:
            attempt = 1
            max_attempts = max_retries
            
            # Track valid links across retries for this question
            accumulated_valid_links = []
            
            # Track previous attempts and feedback for context improvement
            attempt_history = []
            
            # Track if we have a valid answer that just needs links (issue #10)
            validated_answer = None
            skip_answer_checker = False
            
            while attempt <= max_attempts:
                # Step 1: Generate answer
                candidate_answer, doc_urls = self.generate_answer(question, context, char_limit, attempt_history)
                
                if not candidate_answer:
                    return False, "Question Answerer failed to generate an answer", []
                
                # Remove links and citations, save links for documentation
                clean_answer, text_links = self.extract_links_and_clean(candidate_answer)
                
                # Combine documentation URLs from run steps with any URLs found in text
                all_links = list(set(doc_urls + text_links))  # Remove duplicates
                
                # Check character limit
                if len(clean_answer) > char_limit:
                    rejection_reason = f"Answer exceeds character limit ({len(clean_answer)} > {char_limit})"
                    
                    # Add to attempt history for next iteration
                    attempt_history.append({
                        'attempt': attempt,
                        'answer': clean_answer,
                        'rejection_reason': rejection_reason,
                        'rejected_by': 'Character Limit Check'
                    })
                    
                    attempt += 1
                    continue
                
                # Step 2: Validate answer (skip if we already have a validated answer)
                if skip_answer_checker and validated_answer:
                    self.log_reasoning("Answer Checker: Skipping validation - we already have a validated answer")
                    answer_valid = True
                    clean_answer = validated_answer  # Use the previously validated answer
                else:
                    answer_valid, answer_feedback = self.validate_answer(question, clean_answer)
                    
                    if not answer_valid:
                        # Add to attempt history for next iteration
                        attempt_history.append({
                            'attempt': attempt,
                            'answer': clean_answer,
                            'rejection_reason': answer_feedback,
                            'rejected_by': 'Answer Checker'
                        })
                        
                        attempt += 1
                        continue
                    else:
                        # Answer Checker approved - save this as our validated answer
                        validated_answer = clean_answer
                
                # Step 3: Validate links
                links_valid, valid_links, link_feedback = self.validate_links(all_links)
                
                # Accumulate any valid links found in this attempt
                if valid_links:
                    for link in valid_links:
                        if link not in accumulated_valid_links:
                            accumulated_valid_links.append(link)
                
                # Check if we have a valid answer and at least some valid links (current or accumulated)
                if not links_valid and not accumulated_valid_links:
                    # No valid links in current attempt and no accumulated links from previous attempts
                    self.log_reasoning(f"Link Checker rejected: {link_feedback}")
                    
                    # Issue #10 fix: If we have a validated answer but no links, 
                    # don't throw away the answer - instead ask for links that support it
                    if validated_answer and not skip_answer_checker:
                        self.log_reasoning("Answer is validated but has no links")
                        self.log_reasoning("Next attempt will keep the validated answer and ask for supporting links")
                        
                        # Add special context for next iteration to find supporting links
                        attempt_history.append({
                            'attempt': attempt,
                            'answer': validated_answer,
                            'rejection_reason': f"Good answer but needs supporting links: {link_feedback}",
                            'rejected_by': 'Link Checker (needs supporting links)',
                            'special_instruction': 'keep_answer_find_links'
                        })
                        
                        # Set flag to skip Answer Checker on next iteration
                        skip_answer_checker = True
                    else:
                        # Standard rejection - no validated answer yet
                        attempt_history.append({
                            'attempt': attempt,
                            'answer': clean_answer,
                            'rejection_reason': link_feedback,
                            'rejected_by': 'Link Checker'
                        })
                    
                    attempt += 1
                    continue
                elif not links_valid and accumulated_valid_links:
                    # Current attempt has no valid links, but we have accumulated valid links from previous attempts
                    final_valid_links = accumulated_valid_links.copy()
                else:
                    # Current attempt has valid links
                    final_valid_links = accumulated_valid_links.copy()  # Use all accumulated links
                
                # All checks passed
                self.log_reasoning("=" * 80)
                self.log_reasoning("QUESTION WORKFLOW COMPLETED SUCCESSFULLY (Excel Processing)")
                self.log_reasoning("Context will be CLEARED for next question (no carryover of attempt history)")
                self.log_reasoning("=" * 80)
                return True, clean_answer, final_valid_links
                
            # Max attempts reached
            return False, f"Failed to generate acceptable answer after {max_attempts} attempts", []
            
        except Exception as e:
            self.logger.error(f"Error processing question with agents: {e}")
            return False, f"Error: {e}", []
            
    def identify_columns_with_llm_cli(self, df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """Use LLM to identify question, answer, and documentation columns for CLI mode."""
        # Check for mock mode first
        if self.mock_mode:
            return self.identify_columns_mock(df)
            
        try:
            # Create a prompt with column names and sample data
            column_info = []
            for col in df.columns:
                sample_data = df[col].dropna().head(3).tolist()
                # Convert to strings and truncate if too long
                sample_strings = [str(item)[:100] for item in sample_data]
                column_info.append(f"Column '{col}': {sample_strings}")
            
            prompt = f"""Analyze the following Excel columns and classify each column header as one of these types:
- QUESTION: Contains questions to be answered
- RESPONSE: Where AI responses/answers should be written
- DOCUMENTATION: Where reference links/documentation should be written  
- NONE: Not relevant for question answering

Columns:
{chr(10).join(column_info)}

Respond in this exact format:
Question Column: [column_name or NONE if no suitable column found]
Response Column: [column_name or NONE if no suitable column found]  
Documentation Column: [column_name or NONE if no suitable column found]

Only return existing column names. Do not suggest new column names."""

            # Use the Question Answerer agent to analyze columns
            thread = self.project_client.agents.threads.create()
            message = self.project_client.agents.messages.create(
                thread_id=thread.id,
                role="user",
                content=prompt
            )
            
            run = self.project_client.agents.runs.create_and_process(
                thread_id=thread.id,
                agent_id=self.question_answerer_id
            )
            
            messages = self.project_client.agents.messages.list(thread_id=thread.id)
            
            # Parse the response
            for msg in messages:
                if msg.role == "assistant" and msg.content:
                    response = msg.content[0].text.value
                    question_col = self.extract_column_name(response, "Question Column:")
                    answer_col = self.extract_column_name(response, "Response Column:")
                    docs_col = self.extract_column_name(response, "Documentation Column:")
                    
                    # Convert "NONE" to None
                    question_col = None if question_col and question_col.upper() == "NONE" else question_col
                    answer_col = None if answer_col and answer_col.upper() == "NONE" else answer_col
                    docs_col = None if docs_col and docs_col.upper() == "NONE" else docs_col
                    
                    # Validate that columns exist in the dataframe
                    if question_col and question_col not in df.columns:
                        question_col = None
                    if answer_col and answer_col not in df.columns:
                        answer_col = None  
                    if docs_col and docs_col not in df.columns:
                        docs_col = None
                    
                    return question_col, answer_col, docs_col
            
            # Fallback if LLM fails
            return None, None, None
            
        except Exception as e:
            self.logger.error(f"Error identifying columns with LLM: {e}")
            return None, None, None
    
    def extract_column_name(self, response: str, prefix: str) -> Optional[str]:
        """Extract column name from LLM response."""
        lines = response.split('\n')
        for line in lines:
            if line.strip().startswith(prefix):
                # Extract everything after the colon and clean it up
                part = line.split(':', 1)[1].strip()
                # Remove quotes and brackets if present
                part = part.strip('"\'[]')
                return part if part else None
        return None
        
    def run(self):
        """Start the application."""
        if not self.headless_mode:
            self.root.mainloop()
    
    def process_single_question_cli(self, question: str, context: str, char_limit: int, verbose: bool, max_retries: int = None) -> Tuple[bool, str, List[str]]:
        """Process a single question in CLI mode."""
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
        
        if self.tracer:
            with self.tracer.start_as_current_span("questionnaire_agent_cli") as span:
                span.set_attribute("interface.type", "CLI")
                span.set_attribute("question.length", len(question))
                span.set_attribute("question.preview", question[:100])
                span.set_attribute("context", context)
                span.set_attribute("char_limit", char_limit)
                span.set_attribute("verbose_mode", verbose)
                span.set_attribute("max_retries", max_retries)
                
                return self._execute_cli_workflow(question, context, char_limit, verbose, max_retries)
        else:
            return self._execute_cli_workflow(question, context, char_limit, verbose, max_retries)
    
    def _execute_cli_workflow(self, question: str, context: str, char_limit: int, verbose: bool, max_retries: int) -> Tuple[bool, str, List[str]]:
        """Execute the CLI workflow using the same multi-agent approach as GUI."""
        try:
            if verbose:
                print("Starting question processing...")
            
            # Store the current CLI output buffer
            original_cli_output = self.cli_output.copy()
            self.cli_output.clear()
            
            # Create agents if not already created
            if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
                self.create_agents()
            
            # Use the same workflow as GUI mode
            success, answer, links = self._execute_workflow(question, context, char_limit, max_retries)
            
            # Extract results from CLI output (the workflow populates this in headless mode)
            if verbose:
                for line in self.cli_output:
                    print(line)
            
            return success, answer, links
                
        except Exception as e:
            if verbose:
                print(f"Error: {e}")
            return False, f"Error: {e}", []
        finally:
            # Restore original CLI output
            self.cli_output = original_cli_output
    
    
    def process_excel_file_cli(self, input_path: str, output_path: str, context: str, char_limit: int, verbose: bool, max_retries: int = None) -> bool:
        """Process an Excel file in CLI mode."""
        scenario = "questionnaire_agent_excel_processing"
        
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
        
        if self.tracer:
            with self.tracer.start_as_current_span(scenario) as span:
                span.set_attribute("input_file", input_path)
                span.set_attribute("output_file", output_path)
                span.set_attribute("context", context)
                span.set_attribute("char_limit", char_limit)
                span.set_attribute("verbose_mode", verbose)
                span.set_attribute("max_retries", max_retries)
                
                return self._process_excel_file_internal(input_path, output_path, context, char_limit, verbose, span, max_retries)
        else:
            return self._process_excel_file_internal(input_path, output_path, context, char_limit, verbose, None, max_retries)
    
    def _process_excel_file_internal(self, input_path: str, output_path: str, context: str, char_limit: int, verbose: bool, span=None, max_retries: int = None) -> bool:
        """Internal method for Excel file processing with tracing."""
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
            
        try:
            if verbose:
                print(f"Processing Excel file: {input_path}")
            
            # Create agents if not already created
            if not all([self.question_answerer_id, self.answer_checker_id, self.link_checker_id]):
                if span and self.tracer:
                    with self.tracer.start_as_current_span("create_agents") as agent_span:
                        agent_span.set_attribute("operation", "create_all_agents")
                        self.create_agents()
                else:
                    self.create_agents()
            
            # Create temporary file to work with
            import tempfile
            import shutil
            temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            temp_path = temp_file.name
            temp_file.close()
            
            # Copy input file to temporary location
            shutil.copy2(input_path, temp_path)
            if verbose:
                print(f"Created temporary working file: {temp_path}")
            
            # Read Excel file
            # Suppress openpyxl data validation warnings when reading Excel metadata
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                excel_file = pd.ExcelFile(temp_path)
            
            # Load workbook once for all sheets
            from openpyxl import load_workbook
            # Suppress openpyxl data validation warnings when loading Excel
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                wb = load_workbook(temp_path)
            
            total_sheets = len(excel_file.sheet_names)
            if span:
                span.set_attribute("sheets.total_count", total_sheets)
                span.set_attribute("sheets.names", excel_file.sheet_names)
            
            for sheet_index, sheet_name in enumerate(excel_file.sheet_names, 1):
                if verbose:
                    print(f"Processing sheet: {sheet_name}")
                
                if span and self.tracer:
                    with self.tracer.start_as_current_span("process_excel_sheet") as sheet_span:
                        sheet_span.set_attribute("sheet.name", sheet_name)
                        sheet_span.set_attribute("sheet.index", sheet_index)
                        sheet_span.set_attribute("sheet.total", total_sheets)
                        
                        # Suppress openpyxl data validation warnings when reading Excel
                        with warnings.catch_warnings():
                            warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                            df = pd.read_excel(temp_path, sheet_name=sheet_name)
                        sheet_span.set_attribute("sheet.row_count", len(df))
                        sheet_span.set_attribute("sheet.column_count", len(df.columns))
                        
                        self._process_excel_sheet(df, sheet_name, temp_path, wb, context, char_limit, verbose, sheet_span, max_retries)
                else:
                    # Suppress openpyxl data validation warnings when reading Excel
                    with warnings.catch_warnings():
                        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl.worksheet._reader')
                        df = pd.read_excel(temp_path, sheet_name=sheet_name)
                    self._process_excel_sheet(df, sheet_name, temp_path, wb, context, char_limit, verbose, None, max_retries)
            
            # Close the workbook after processing all sheets
            wb.close()
            
            # Copy the completed temporary file to the final output location
            shutil.copy2(temp_path, output_path)
            if verbose:
                print(f"Copied completed file from {temp_path} to {output_path}")
            
            # Clean up temporary file
            try:
                import os
                os.unlink(temp_path)
                if verbose:
                    print(f"Cleaned up temporary file: {temp_path}")
            except Exception as cleanup_error:
                self.logger.warning(f"Could not clean up temporary file {temp_path}: {cleanup_error}")
            
            if verbose:
                print(f"Excel processing completed. Results saved to: {output_path}")
            
            if span:
                span.set_attribute("success", True)
            
            return True
                
        except Exception as e:
            error_msg = f"Error processing Excel file: {e}"
            self.logger.error(error_msg)
            if verbose:
                print(error_msg)
            
            # Clean up temporary file if it exists
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    import os
                    os.unlink(temp_path)
                    if verbose:
                        print(f"Cleaned up temporary file after error: {temp_path}")
                except Exception as cleanup_error:
                    self.logger.warning(f"Could not clean up temporary file {temp_path} after error: {cleanup_error}")
            
            if span:
                span.set_attribute("success", False)
                span.set_attribute("error.message", str(e))
                span.set_attribute("error.type", type(e).__name__)
                span.set_status(trace.Status(trace.StatusCode.ERROR, str(e)))
            
            return False
    
    def _process_excel_sheet(self, df: pd.DataFrame, sheet_name: str, temp_path: str, wb, context: str, char_limit: int, verbose: bool, span=None, max_retries: int = None):
        """Process a single Excel sheet with tracing."""
        # Use instance default if not provided
        if max_retries is None:
            max_retries = self.max_retries
            
        # Use LLM to identify columns
        question_col, answer_col, docs_col = self.identify_columns_with_llm_cli(df)
        
        if verbose:
            print(f"LLM identified columns - Questions: {question_col}, Answers: {answer_col}, Docs: {docs_col}")
        
        if span:
            span.set_attribute("columns.question", question_col or "none")
            span.set_attribute("columns.answer", answer_col or "none") 
            span.set_attribute("columns.docs", docs_col or "none")
        
        # Skip sheet if no question or answer column found
        if not question_col or not answer_col:
            if verbose:
                print(f"Skipping sheet '{sheet_name}' - missing required question or answer column")
            if span:
                span.set_attribute("sheet.skipped", True)
                span.set_attribute("skip_reason", "missing_required_columns")
            return

        # Get worksheet for this sheet
        ws = wb[sheet_name]
        
        # Convert answer and documentation columns to object dtype to prevent pandas dtype warnings
        if answer_col in df.columns:
            df[answer_col] = df[answer_col].astype('object')
        if docs_col and docs_col in df.columns:
            df[docs_col] = df[docs_col].astype('object')
        
        # Process each question
        questions_processed = 0
        questions_attempted = 0
        for idx, row in df.iterrows():
            if pd.notna(row[question_col]) and str(row[question_col]).strip():
                question = str(row[question_col]).strip()
                questions_attempted += 1
                if verbose:
                    print(f"Processing question {idx + 1}: {question[:50]}...")
                
                # Update the Question box to show current question being processed (UI mode only)
                if not self.headless_mode:
                    self.root.after(0, lambda q=question: self.update_question_display(q))
                
                # Process question using CLI workflow
                success, answer, links = self.process_single_question_cli(question, context, char_limit, False, max_retries)
                
                if success:
                    # Show answer preview and link count in verbose mode
                    if verbose:
                        # Strip Azure AI context prefix from answer preview
                        clean_answer = answer
                        # Look for pattern "Based on ... context, regarding '...': " and strip it
                        import re
                        context_pattern = r"^Based on .+ context, regarding '.+': "
                        match = re.search(context_pattern, answer)
                        if match:
                            clean_answer = answer[match.end():]
                        
                        # Show beginning of the clean answer (first 100 characters)
                        answer_preview = clean_answer[:100] + "..." if len(clean_answer) > 100 else clean_answer
                        print(f"Answer: {answer_preview}")
                        # Show how many links were found
                        link_count = len(links) if links else 0
                        print(f"Found {link_count} documentation links")
                    
                    # Update the answer column in dataframe
                    df.at[idx, answer_col] = answer
                    
                    # Update documentation column only if it exists and we have links
                    if docs_col and links:
                        df.at[idx, docs_col] = '\n'.join(links)
                    
                    # Save to Excel file immediately using openpyxl
                    row_num = idx + 2  # +2 because Excel is 1-indexed and has header
                    
                    # Find and update answer column
                    for col_idx, col_name in enumerate(df.columns, 1):
                        if col_name == answer_col:
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.value = answer
                        elif col_name == docs_col and docs_col and links:
                            cell = ws.cell(row=row_num, column=col_idx)
                            cell.value = '\n'.join(links)
                    
                    # Save the workbook immediately after each successful question
                    wb.save(temp_path)
                    
                    if verbose:
                        print(f"Successfully processed question {idx + 1} and saved to temporary file")
                    questions_processed += 1
                else:
                    if verbose:
                        print(f"Failed to process question {idx + 1}: {answer}")
                    # Leave response blank on failure - don't write error messages
                    # Leave documentation blank on failure - don't write error messages
        
        if span:
            span.set_attribute("questions.attempted", questions_attempted)
            span.set_attribute("questions.processed", questions_processed)
            span.set_attribute("questions.success_rate", questions_processed / questions_attempted if questions_attempted > 0 else 0)
        
        if verbose:
            print(f"Processed {questions_processed}/{questions_attempted} questions successfully in sheet '{sheet_name}'")


def create_cli_parser():
    """Create command line argument parser."""
    parser = argparse.ArgumentParser(
        description="Questionnaire Multiagent - AI question answering with fact-checking and link validation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""Examples:
  # Single question
  python question_answerer.py --question "Does your service offer video generative AI?" --context "Microsoft Azure AI" --char-limit 2000
  
  # Excel processing
  python question_answerer.py --import-excel questions.xlsx --output-excel processed.xlsx --context "Microsoft Azure AI" --verbose"""
    )
    
    parser.add_argument('-q', '--question', type=str, help='The natural-language question to ask')
    parser.add_argument('-c', '--context', type=str, default='Microsoft Azure AI', help='Context or topic string to bias the question answering (default: "Microsoft Azure AI")')
    parser.add_argument('--char-limit', type=int, default=2000, help='Integer character limit for the final answer (default: 2000)')
    parser.add_argument('--max-retries', type=int, default=10, help='Maximum number of retries for answer generation (default: 10)')
    parser.add_argument('--import-excel', type=str, metavar='PATH', help='Path to an Excel file to process in batch')
    parser.add_argument('--output-excel', type=str, metavar='PATH', help='Path where the processed Excel file will be written')
    parser.add_argument('--verbose', action='store_true', default=True, help='Enable verbose/reasoning log output (default: True)')
    parser.add_argument('--mock', action='store_true', help='Enable mock mode for testing (does not require Azure credentials)')
    
    return parser


def main():
    """Main entry point."""
    parser = create_cli_parser()
    
    # Check if any arguments were provided
    if len(sys.argv) == 1:
        # No arguments - run in GUI mode
        app = None
        try:
            app = QuestionnaireAgentUI(headless_mode=False)
            app.run()
        except Exception as e:
            print(f"Failed to start application: {e}")
            try:
                messagebox.showerror("Startup Error", f"Failed to start application:\n{e}")
            except:
                pass
        finally:
            # Ensure cleanup happens even if app creation failed
            if app:
                app.cleanup_agents()
    else:
        # Arguments provided - run in CLI mode
        args = parser.parse_args()
        
        # Validate arguments
        if not args.question and not args.import_excel:
            print("Error: Either --question or --import-excel must be provided")
            parser.print_help()
            sys.exit(1)
        
        if args.import_excel and not args.output_excel:
            # Generate default output filename
            input_path = Path(args.import_excel)
            args.output_excel = str(input_path.parent / f"{input_path.stem}.answered.xlsx")
        
        app = None
        try:
            app = QuestionnaireAgentUI(headless_mode=True, max_retries=args.max_retries, mock_mode=args.mock)
            
            if args.question:
                # Process single question
                success, answer, links = app.process_single_question_cli(
                    args.question, args.context, args.char_limit, args.verbose, args.max_retries
                )
                
                if success:
                    print("\n=== ANSWER ===")
                    print(answer)
                    if links:
                        print("\n=== DOCUMENTATION LINKS ===")
                        for link in links:
                            print(f"• {link}")
                    else:
                        print("\n=== DOCUMENTATION LINKS ===")
                        print("No documentation links found")
                    # Cleanup before exit
                    if app:
                        app.cleanup_agents()
                    sys.exit(0)
                else:
                    print(f"Error: {answer}")
                    # Cleanup before exit
                    if app:
                        app.cleanup_agents()
                    sys.exit(1)
            
            elif args.import_excel:
                # Process Excel file
                if not os.path.exists(args.import_excel):
                    print(f"Error: Excel file not found: {args.import_excel}")
                    # Cleanup before exit
                    if app:
                        app.cleanup_agents()
                    sys.exit(1)
                
                success = app.process_excel_file_cli(
                    args.import_excel, args.output_excel, args.context, args.char_limit, args.verbose, args.max_retries
                )
                
                if success:
                    print(f"\nExcel processing completed successfully. Results saved to: {args.output_excel}")
                    # Cleanup before exit
                    if app:
                        app.cleanup_agents()
                    sys.exit(0)
                else:
                    print("Error: Excel processing failed")
                    # Cleanup before exit
                    if app:
                        app.cleanup_agents()
                    sys.exit(1)
        
        except KeyboardInterrupt:
            print("\nOperation cancelled by user")
            # Cleanup before exit
            if app:
                app.cleanup_agents()
            sys.exit(1)
        except Exception as e:
            print(f"Error: {e}")
            # Cleanup before exit
            if app:
                app.cleanup_agents()
            sys.exit(1)


if __name__ == "__main__":
    main()