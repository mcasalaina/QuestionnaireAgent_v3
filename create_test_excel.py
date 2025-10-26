#!/usr/bin/env python3
"""Create a simple Excel test file for testing the import functionality."""

import sys
import os

# Add src to path
src_path = os.path.join(os.path.dirname(__file__), 'src')
sys.path.insert(0, src_path)

print("Creating test Excel file...")

import openpyxl
from openpyxl.workbook import Workbook

# Create test Excel file with sample questions
wb = Workbook()
ws = wb.active
ws.title = 'Questions'

# Add header
ws['A1'] = 'Question'
ws['B1'] = 'Response'

# Add sample questions (simple ones to test quickly)
questions = [
    'What is Python?',
    'What is Microsoft Azure?',
    'What is artificial intelligence?'
]

for i, question in enumerate(questions, start=2):
    ws[f'A{i}'] = question

# Save the test file
wb.save('test_simple_questions.xlsx')
wb.close()
print('Created test_simple_questions.xlsx with 3 simple questions')

# Verify file exists
if os.path.exists('test_simple_questions.xlsx'):
    print('✓ File created successfully')
else:
    print('✗ File creation failed')