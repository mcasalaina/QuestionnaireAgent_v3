"""Script to create test Excel files for live Excel processing."""

import openpyxl
from openpyxl.workbook import Workbook
import os

def create_single_sheet_file():
    """Create single_sheet_5_questions.xlsx with 5 questions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    
    # Headers
    ws['A1'] = "Question"
    ws['B1'] = "Response"
    
    # Questions
    questions = [
        "What is artificial intelligence?",
        "How does machine learning work?",
        "What are the benefits of cloud computing?",
        "Explain natural language processing.",
        "What is the difference between AI and ML?"
    ]
    
    for i, question in enumerate(questions, start=2):
        ws[f'A{i}'] = question
    
    return wb

def create_multi_sheet_file():
    """Create multi_sheet_3x10_questions.xlsx with 3 sheets, 10 questions each."""
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    sheet_topics = [
        ("AI Basics", [
            "What is artificial intelligence?",
            "Define machine learning.",
            "What is deep learning?",
            "How do neural networks work?",
            "What is computer vision?",
            "Define natural language processing.",
            "What is reinforcement learning?",
            "How does supervised learning work?",
            "Explain unsupervised learning.",
            "What are the types of AI?"
        ]),
        ("Cloud Computing", [
            "What is cloud computing?",
            "Define Infrastructure as a Service.",
            "What is Platform as a Service?",
            "Explain Software as a Service.",
            "What are the benefits of cloud?",
            "What is Azure?",
            "Define AWS.",
            "What is Google Cloud Platform?",
            "Explain hybrid cloud.",
            "What is multi-cloud strategy?"
        ]),
        ("Data Science", [
            "What is data science?",
            "Define big data.",
            "What is data mining?",
            "Explain data visualization.",
            "What is statistical analysis?",
            "Define predictive analytics.",
            "What is data preprocessing?",
            "Explain feature engineering.",
            "What is model validation?",
            "Define cross-validation."
        ])
    ]
    
    for sheet_name, questions in sheet_topics:
        ws = wb.create_sheet(title=sheet_name)
        
        # Headers
        ws['A1'] = "Question"
        ws['B1'] = "Response"
        
        # Questions
        for i, question in enumerate(questions, start=2):
            ws[f'A{i}'] = question
    
    return wb

def create_hidden_sheets_file():
    """Create hidden_sheets.xlsx with visible and hidden sheets."""
    wb = Workbook()
    
    # First sheet - visible
    ws1 = wb.active
    ws1.title = "Visible Questions"
    ws1['A1'] = "Question"
    ws1['B1'] = "Response"
    
    visible_questions = [
        "What is Azure AI?",
        "How to use Azure OpenAI?",
        "What is Cognitive Services?"
    ]
    
    for i, question in enumerate(visible_questions, start=2):
        ws1[f'A{i}'] = question
    
    # Second sheet - hidden with data
    ws2 = wb.create_sheet(title="Hidden Config")
    ws2.sheet_state = 'hidden'
    ws2['A1'] = "Configuration"
    ws2['A2'] = "Secret API Key"
    ws2['A3'] = "Internal Settings"
    
    # Third sheet - visible
    ws3 = wb.create_sheet(title="More Questions")
    ws3['A1'] = "Question"
    ws3['B1'] = "Response"
    
    more_questions = [
        "What is machine learning in Azure?",
        "How to deploy AI models?"
    ]
    
    for i, question in enumerate(more_questions, start=2):
        ws3[f'A{i}'] = question
    
    return wb

def create_invalid_format_file():
    """Create invalid_format.xlsx - corrupted/empty file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    # No questions - just empty sheet
    return wb

def main():
    """Create all test Excel files."""
    fixtures_dir = "C:\\src\\QuestionnaireAgent_v3\\tests\\fixtures\\excel"
    
    # Create single sheet file
    wb1 = create_single_sheet_file()
    wb1.save(os.path.join(fixtures_dir, "single_sheet_5_questions.xlsx"))
    wb1.close()
    print("Created single_sheet_5_questions.xlsx")
    
    # Create multi sheet file
    wb2 = create_multi_sheet_file()
    wb2.save(os.path.join(fixtures_dir, "multi_sheet_3x10_questions.xlsx"))
    wb2.close()
    print("Created multi_sheet_3x10_questions.xlsx")
    
    # Create hidden sheets file
    wb3 = create_hidden_sheets_file()
    wb3.save(os.path.join(fixtures_dir, "hidden_sheets.xlsx"))
    wb3.close()
    print("Created hidden_sheets.xlsx")
    
    # Create invalid format file
    wb4 = create_invalid_format_file()
    wb4.save(os.path.join(fixtures_dir, "invalid_format.xlsx"))
    wb4.close()
    print("Created invalid_format.xlsx")
    
    print("All test Excel files created successfully!")

if __name__ == "__main__":
    main()