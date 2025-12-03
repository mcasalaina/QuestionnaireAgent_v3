"""Unit tests for Excel blank row and section header handling."""

import pytest
import sys
import os

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'src'))

from excel.loader import ExcelLoader


class TestSectionHeaderDetection:
    """Test heuristic detection of section headers."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.loader = ExcelLoader()
    
    def test_detect_section_prefix(self):
        """Test detection of common section prefixes."""
        assert self.loader._is_section_header("Section 1: Introduction") is True
        assert self.loader._is_section_header("Part A: Basic Questions") is True
        assert self.loader._is_section_header("Chapter 3: Advanced Topics") is True
        assert self.loader._is_section_header("Category: General") is True
        assert self.loader._is_section_header("Topic 5: Cloud Services") is True
    
    def test_detect_decorated_headers(self):
        """Test detection of headers wrapped in decorators."""
        assert self.loader._is_section_header("--- Advanced Topics ---") is True
        assert self.loader._is_section_header("=== Section 2 ===") is True
        assert self.loader._is_section_header("*** Important ***") is True
        assert self.loader._is_section_header("### Headers ###") is True
    
    def test_detect_uppercase_headers(self):
        """Test detection of all-uppercase short headers."""
        assert self.loader._is_section_header("ADVANCED TOPICS") is True
        assert self.loader._is_section_header("SECTION TWO") is True
        # Long uppercase text might be a question
        assert self.loader._is_section_header("A" * 50) is True  # Still detected as header (all caps)
    
    def test_allow_valid_questions(self):
        """Test that valid questions are not detected as headers."""
        assert self.loader._is_section_header("What is Azure?") is False
        assert self.loader._is_section_header("How does machine learning work?") is False
        assert self.loader._is_section_header("Can you explain the concept of containerization?") is False
        assert self.loader._is_section_header("What are the benefits of cloud computing?") is False
        assert self.loader._is_section_header("Why should I use Azure Cognitive Services?") is False
    
    def test_short_text_detected(self):
        """Test that very short text is detected as potential header."""
        assert self.loader._is_section_header("A") is True
        assert self.loader._is_section_header("OK") is True
        assert self.loader._is_section_header("") is True
    
    def test_numbered_sections(self):
        """Test detection of numbered section headers."""
        assert self.loader._is_section_header("1. Introduction") is True
        assert self.loader._is_section_header("1.1 Overview") is True
        assert self.loader._is_section_header("I. First Section") is True
        # Numbered questions should NOT be detected as headers
        assert self.loader._is_section_header("1. What is your question?") is False


class TestBlankRowHandling:
    """Test blank row handling in Excel loading."""
    
    def setup_method(self):
        """Set up test fixtures."""
        self.loader = ExcelLoader()
        self.temp_dir = '/tmp/test_excel_blank_rows'
        os.makedirs(self.temp_dir, exist_ok=True)
    
    def test_load_with_blank_rows(self):
        """Test loading Excel file with blank rows."""
        import openpyxl
        from openpyxl.workbook import Workbook
        
        # Create test file
        wb = Workbook()
        ws = wb.active
        ws.title = 'Questions'
        ws['A1'] = 'Question'
        ws['B1'] = 'Response'
        
        ws['A2'] = 'Question 1?'
        ws['A3'] = 'Question 2?'
        # Row 4 is blank
        ws['A5'] = 'Question 3?'
        # Row 6 is blank
        ws['A7'] = 'Question 4?'
        
        file_path = os.path.join(self.temp_dir, 'blank_rows_test.xlsx')
        wb.save(file_path)
        wb.close()
        
        # Load and verify
        workbook_data = self.loader.load_workbook(file_path)
        assert len(workbook_data.sheets) == 1
        
        sheet = workbook_data.sheets[0]
        assert len(sheet.questions) == 4
        assert sheet.questions == [
            'Question 1?',
            'Question 2?',
            'Question 3?',
            'Question 4?'
        ]
        assert sheet.row_indices == [2, 3, 5, 7]
    
    def test_load_with_section_headers(self):
        """Test loading Excel file with section headers."""
        import openpyxl
        from openpyxl.workbook import Workbook
        
        # Create test file
        wb = Workbook()
        ws = wb.active
        ws.title = 'Questions'
        ws['A1'] = 'Question'
        ws['B1'] = 'Response'
        
        ws['A2'] = 'Question 1?'
        ws['A3'] = 'Section 2: Advanced Topics'  # Section header
        ws['A4'] = 'Question 2?'
        ws['A5'] = '--- Category ---'  # Decorated header
        ws['A6'] = 'Question 3?'
        
        file_path = os.path.join(self.temp_dir, 'section_headers_test.xlsx')
        wb.save(file_path)
        wb.close()
        
        # Load and verify
        workbook_data = self.loader.load_workbook(file_path)
        sheet = workbook_data.sheets[0]
        
        assert len(sheet.questions) == 3
        assert sheet.questions == [
            'Question 1?',
            'Question 2?',
            'Question 3?'
        ]
        assert sheet.row_indices == [2, 4, 6]
    
    def test_save_with_row_indices(self):
        """Test saving answers to correct rows when using row indices."""
        import openpyxl
        from openpyxl.workbook import Workbook
        
        # Create test file
        wb = Workbook()
        ws = wb.active
        ws.title = 'Questions'
        ws['A1'] = 'Question'
        ws['B1'] = 'Response'
        
        ws['A2'] = 'Question 1?'
        # Row 3 blank
        ws['A4'] = 'Section header'
        ws['A5'] = 'Question 2?'
        
        file_path = os.path.join(self.temp_dir, 'save_rows_test.xlsx')
        wb.save(file_path)
        wb.close()
        
        # Load, add answers, and save
        workbook_data = self.loader.load_workbook(file_path)
        sheet = workbook_data.sheets[0]
        sheet.answers[0] = 'Answer 1'
        sheet.answers[1] = 'Answer 2'
        
        output_path = os.path.join(self.temp_dir, 'save_rows_output.xlsx')
        self.loader.save_workbook(workbook_data, output_path)
        
        # Verify output
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        assert ws.cell(row=2, column=2).value == 'Answer 1'
        assert ws.cell(row=3, column=2).value is None  # Blank row
        assert ws.cell(row=4, column=2).value is None  # Section header row
        assert ws.cell(row=5, column=2).value == 'Answer 2'
        wb.close()


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
