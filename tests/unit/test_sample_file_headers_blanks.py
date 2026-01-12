"""Test the sample_questionnaire_simple_headers_blanks.xlsx file handling."""

import pytest
import sys
import os
import openpyxl

# Add src to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', 'src'))

from excel.loader import ExcelLoader


class TestSampleFileHeadersBlanks:
    """Test handling of sample file with section headers and blank rows."""

    def setup_method(self):
        """Set up test fixtures."""
        self.loader = ExcelLoader()
        self.test_file = os.path.join(
            os.path.dirname(__file__),
            '..',
            'sample_questionnaire_simple_headers_blanks.xlsx'
        )
        self.output_dir = '/tmp/test_sample_headers_blanks'
        os.makedirs(self.output_dir, exist_ok=True)

    def test_load_sample_file(self):
        """Test that sample file loads correctly with headers and blanks skipped."""
        workbook_data = self.loader.load_workbook(self.test_file)

        # Should have 1 sheet
        assert len(workbook_data.sheets) == 1

        sheet = workbook_data.sheets[0]

        # Should have loaded 9 questions (skipping section headers and blanks)
        assert len(sheet.questions) == 9

        # Verify expected questions
        expected_questions = [
            'What is your full name?',
            'What is your email?',
            'Describe your experience with cloud computing',
            'What programming languages are you familiar with?',
            'What specific Azure services have you worked with?',
            'Have you implemented any AI/ML solutions using Azure?',
            'Are there any other technical skills or certifications you would like to mention?',
            'What are your availability and preferred working hours?',
            'Please provide any additional comments or questions'
        ]
        assert sheet.questions == expected_questions

        # Verify row indices are correct (skipping headers and blanks)
        expected_row_indices = [4, 5, 9, 10, 14, 15, 19, 20, 21]
        assert sheet.row_indices == expected_row_indices

    def test_save_preserves_headers_and_blanks(self):
        """Test that saving responses doesn't overwrite section headers or blank rows."""
        # Load file
        workbook_data = self.loader.load_workbook(self.test_file)
        sheet = workbook_data.sheets[0]

        # Add answers to all questions
        answers = [
            'John Doe',
            'john.doe@example.com',
            '5 years of experience with Azure, AWS, and GCP',
            'Python, JavaScript, TypeScript, Go',
            'Azure Functions, App Service, Cosmos DB, Cognitive Services',
            'Yes, built a chatbot using Azure OpenAI',
            'Azure Solutions Architect Expert certified',
            'Monday-Friday 9am-5pm EST',
            'Looking forward to contributing!'
        ]
        sheet.answers = answers

        # Save to output file
        output_path = os.path.join(self.output_dir, 'output_with_answers.xlsx')
        self.loader.save_workbook(workbook_data, output_path)

        # Verify output file structure
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Row 1: Headers should be unchanged
        assert ws.cell(row=1, column=1).value == 'Question'
        assert ws.cell(row=1, column=2).value == 'Response'

        # Row 2: Section header should be unchanged and no response written
        assert ws.cell(row=2, column=1).value == 'Section 1: Basic Information'
        assert ws.cell(row=2, column=2).value is None

        # Row 3: Blank row should remain blank
        assert ws.cell(row=3, column=1).value is None
        assert ws.cell(row=3, column=2).value is None

        # Row 4-5: Questions with answers
        assert ws.cell(row=4, column=1).value == 'What is your full name?'
        assert ws.cell(row=4, column=2).value == 'John Doe'
        assert ws.cell(row=5, column=1).value == 'What is your email?'
        assert ws.cell(row=5, column=2).value == 'john.doe@example.com'

        # Row 6: Blank row should remain blank
        assert ws.cell(row=6, column=1).value is None
        assert ws.cell(row=6, column=2).value is None

        # Row 7: Section header should be unchanged and no response written
        assert ws.cell(row=7, column=1).value == 'Section 2: Technical Background and Experience'
        assert ws.cell(row=7, column=2).value is None

        # Row 8: Blank row should remain blank
        assert ws.cell(row=8, column=1).value is None
        assert ws.cell(row=8, column=2).value is None

        # Row 9-10: Questions with answers
        assert ws.cell(row=9, column=1).value == 'Describe your experience with cloud computing'
        assert ws.cell(row=9, column=2).value == '5 years of experience with Azure, AWS, and GCP'
        assert ws.cell(row=10, column=1).value == 'What programming languages are you familiar with?'
        assert ws.cell(row=10, column=2).value == 'Python, JavaScript, TypeScript, Go'

        # Row 11: Blank row should remain blank
        assert ws.cell(row=11, column=1).value is None
        assert ws.cell(row=11, column=2).value is None

        # Row 12: Section header should be unchanged and no response written
        assert ws.cell(row=12, column=1).value == 'Section 3: Project-Specific Questions - Azure AI'
        assert ws.cell(row=12, column=2).value is None

        # Row 13: Blank row should remain blank
        assert ws.cell(row=13, column=1).value is None
        assert ws.cell(row=13, column=2).value is None

        # Row 14-15: Questions with answers
        assert ws.cell(row=14, column=1).value == 'What specific Azure services have you worked with?'
        assert ws.cell(row=14, column=2).value == 'Azure Functions, App Service, Cosmos DB, Cognitive Services'
        assert ws.cell(row=15, column=1).value == 'Have you implemented any AI/ML solutions using Azure?'
        assert ws.cell(row=15, column=2).value == 'Yes, built a chatbot using Azure OpenAI'

        # Row 16: Blank row should remain blank
        assert ws.cell(row=16, column=1).value is None
        assert ws.cell(row=16, column=2).value is None

        # Row 17: Section header (decorated) should be unchanged and no response written
        assert ws.cell(row=17, column=1).value == '--- Section 4: Additional Details ---'
        assert ws.cell(row=17, column=2).value is None

        # Row 18: Blank row should remain blank
        assert ws.cell(row=18, column=1).value is None
        assert ws.cell(row=18, column=2).value is None

        # Row 19-21: Questions with answers
        assert ws.cell(row=19, column=1).value == 'Are there any other technical skills or certifications you would like to mention?'
        assert ws.cell(row=19, column=2).value == 'Azure Solutions Architect Expert certified'
        assert ws.cell(row=20, column=1).value == 'What are your availability and preferred working hours?'
        assert ws.cell(row=20, column=2).value == 'Monday-Friday 9am-5pm EST'
        assert ws.cell(row=21, column=1).value == 'Please provide any additional comments or questions'
        assert ws.cell(row=21, column=2).value == 'Looking forward to contributing!'

        wb.close()


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
